from flask import (
    Flask,
    render_template,
    request,
    redirect,
    url_for,
    send_file,
    send_from_directory,
    session,
    g,
    jsonify,
)
import sqlite3
import re
import csv
import os
import json
import base64
import hashlib
import binascii
import smtplib
import ssl
import unicodedata
from email.message import EmailMessage
from urllib.parse import urlencode, urlparse
from urllib.request import Request, urlopen
from urllib.error import URLError, HTTPError
from uuid import uuid4
from contextlib import closing
from datetime import datetime, timedelta
from io import BytesIO
from secrets import token_urlsafe

from werkzeug.security import check_password_hash, generate_password_hash

try:
    from cryptography.fernet import Fernet, InvalidToken
except Exception:
    Fernet = None
    InvalidToken = Exception

try:
    import openpyxl
except Exception:
    openpyxl = None

try:
    from PIL import Image, ImageOps
except Exception:
    Image = None
    ImageOps = None

APP_TITLE = "Solicitação de compras"
DB_PATH = (os.environ.get("DB_PATH") or "app.db").strip()
LOGO_FILE_BASENAME = "logo_csc"
LOGO_EXTENSIONS = {".png", ".jpg", ".jpeg"}
IMAGE_EXTENSIONS = {".png", ".jpg", ".jpeg", ".webp", ".jfif"}
REQUISICAO_IMAGE_EXTENSIONS = IMAGE_EXTENSIONS
MAX_IMAGE_SIZE_BYTES = 8 * 1024 * 1024
MAX_LOGO_SIZE_BYTES = 12 * 1024 * 1024
REQUISICAO_IMAGE_MAX_SIDE_PX = 1800
REQUEST_LEAD_RULES = [
    (10, 7),
    (30, 14),
    (50, 18),
    (999, 21),
]
REQUEST_LEAD_FALLBACK_DAYS = 21
GOOGLE_DISCOVERY_URL = os.environ.get(
    "GOOGLE_DISCOVERY_URL",
    "https://accounts.google.com/.well-known/openid-configuration",
).strip()
GOOGLE_CLIENT_ID = (os.environ.get("GOOGLE_CLIENT_ID") or "").strip()
GOOGLE_CLIENT_SECRET = (os.environ.get("GOOGLE_CLIENT_SECRET") or "").strip()
MASTER_ADMIN_NAME = (os.environ.get("MASTER_ADMIN_NAME") or "Administrador Master").strip()
MASTER_ADMIN_EMAIL = (
    os.environ.get("MASTER_ADMIN_EMAIL") or "mangefestekaua@gmail.com"
).strip()
MASTER_ADMIN_PASSWORD = (
    os.environ.get("MASTER_ADMIN_PASSWORD") or "14012005kK@"
).strip()
ADMIN_PASSWORD_VIEW_KEY = (os.environ.get("ADMIN_PASSWORD_VIEW_KEY") or "").strip()

_password_cipher = None
_password_cipher_ready = False


def env_flag(name: str, default: bool = False) -> bool:
    value = os.environ.get(name)
    if value is None:
        return default
    return value.strip().lower() in {"1", "true", "yes", "on", "y"}


def env_int(name: str, default: int) -> int:
    value = os.environ.get(name)
    if value is None:
        return default
    try:
        return int(value.strip())
    except (TypeError, ValueError):
        return default


AUTH_ALLOW_PUBLIC_SIGNUP = env_flag("AUTH_ALLOW_PUBLIC_SIGNUP", True)
USER_ONLINE_WINDOW_SECONDS = max(30, env_int("USER_ONLINE_WINDOW_SECONDS", 180))
USER_LAST_SEEN_WRITE_INTERVAL_SECONDS = max(
    15, env_int("USER_LAST_SEEN_WRITE_INTERVAL_SECONDS", 60)
)
PASSWORD_RESET_TOKEN_MINUTES = max(15, env_int("PASSWORD_RESET_TOKEN_MINUTES", 30))
SMTP_HOST = (os.environ.get("SMTP_HOST") or "").strip()
SMTP_PORT = env_int("SMTP_PORT", 587)
SMTP_USERNAME = (os.environ.get("SMTP_USERNAME") or "").strip()
SMTP_PASSWORD = (os.environ.get("SMTP_PASSWORD") or "").strip()
SMTP_FROM_EMAIL = (os.environ.get("SMTP_FROM_EMAIL") or SMTP_USERNAME).strip()
SMTP_USE_TLS = env_flag("SMTP_USE_TLS", True)
SMTP_USE_SSL = env_flag("SMTP_USE_SSL", False)
PUBLIC_ENDPOINTS = {
    "login",
    "signup",
    "logout",
    "forgot_password",
    "reset_password",
    "login_google",
    "google_callback",
    "favicon",
}

app = Flask(__name__)
app.config["SECRET_KEY"] = (os.environ.get("FLASK_SECRET_KEY") or "csc-change-me").strip()
app.config["SESSION_COOKIE_HTTPONLY"] = True
app.config["SESSION_COOKIE_SAMESITE"] = "Lax"
app.config["SESSION_COOKIE_SECURE"] = env_flag("SESSION_COOKIE_SECURE", False)

_google_openid_cache = None


@app.route("/favicon.ico")
def favicon():
    return send_from_directory(
        os.path.join(app.root_path, "static"),
        "favicon.ico",
        mimetype="image/x-icon",
    )


def get_db():
    db_dir = os.path.dirname(os.path.abspath(DB_PATH))
    if db_dir and not os.path.exists(db_dir):
        os.makedirs(db_dir, exist_ok=True)
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    return conn


def now_iso() -> str:
    return datetime.now().isoformat()


def parse_iso_datetime(value: str):
    raw = (value or "").strip()
    if not raw:
        return None
    candidate = raw.replace("Z", "+00:00")
    try:
        dt = datetime.fromisoformat(candidate)
    except ValueError:
        return None
    if dt.tzinfo is not None:
        dt = dt.astimezone().replace(tzinfo=None)
    return dt


def humanize_elapsed(delta: timedelta) -> str:
    total_seconds = int(max(0, delta.total_seconds()))
    if total_seconds < 30:
        return "agora"
    if total_seconds < 60:
        return f"{total_seconds}s"
    if total_seconds < 3600:
        minutes = total_seconds // 60
        return f"{minutes} min"
    if total_seconds < 86400:
        hours = total_seconds // 3600
        minutes = (total_seconds % 3600) // 60
        if minutes == 0:
            return f"{hours} h"
        return f"{hours} h {minutes} min"
    days = total_seconds // 86400
    hours = (total_seconds % 86400) // 3600
    if hours == 0:
        return f"{days} d"
    return f"{days} d {hours} h"


def format_dt_pt(dt):
    if not dt:
        return "-"
    return dt.strftime("%d/%m/%Y %H:%M")


def normalize_email(email: str) -> str:
    return (email or "").strip().lower()


def is_password_reset_email_ready() -> bool:
    return bool(SMTP_HOST and SMTP_PORT > 0 and SMTP_FROM_EMAIL)


def build_reset_token_hash(raw_token: str) -> str:
    token = (raw_token or "").strip()
    if not token:
        return ""
    return hashlib.sha256(token.encode("utf-8")).hexdigest()


def send_password_reset_email(recipient_email: str, recipient_name: str, reset_link: str) -> bool:
    to_email = normalize_email(recipient_email)
    if not to_email or not reset_link or not is_password_reset_email_ready():
        return False

    subject = f"{APP_TITLE} - Redefinicao de senha"
    nome = (recipient_name or "").strip() or to_email
    body_text = (
        f"Ola, {nome}.\n\n"
        "Recebemos uma solicitacao para redefinir sua senha.\n"
        f"Use este link para continuar: {reset_link}\n\n"
        f"Este link expira em {PASSWORD_RESET_TOKEN_MINUTES} minutos.\n"
        "Se voce nao solicitou a redefinicao, ignore este email.\n"
    )
    body_html = (
        "<p>Ola, {nome}.</p>"
        "<p>Recebemos uma solicitacao para redefinir sua senha.</p>"
        f"<p><a href=\"{reset_link}\">Clique aqui para redefinir sua senha</a></p>"
        f"<p>Este link expira em {PASSWORD_RESET_TOKEN_MINUTES} minutos.</p>"
        "<p>Se voce nao solicitou a redefinicao, ignore este email.</p>"
    ).format(nome=nome)

    message = EmailMessage()
    message["Subject"] = subject
    message["From"] = SMTP_FROM_EMAIL
    message["To"] = to_email
    message.set_content(body_text)
    message.add_alternative(body_html, subtype="html")

    try:
        if SMTP_USE_SSL:
            with smtplib.SMTP_SSL(SMTP_HOST, SMTP_PORT, timeout=15) as smtp:
                if SMTP_USERNAME:
                    smtp.login(SMTP_USERNAME, SMTP_PASSWORD)
                smtp.send_message(message)
        else:
            with smtplib.SMTP(SMTP_HOST, SMTP_PORT, timeout=15) as smtp:
                if SMTP_USE_TLS:
                    smtp.starttls(context=ssl.create_default_context())
                if SMTP_USERNAME:
                    smtp.login(SMTP_USERNAME, SMTP_PASSWORD)
                smtp.send_message(message)
        return True
    except Exception as exc:
        app.logger.error("Falha ao enviar email de redefinicao: %s", exc)
        return False


def fetch_valid_password_reset_token(conn, raw_token: str):
    token_hash = build_reset_token_hash(raw_token)
    if not token_hash:
        return None
    row = conn.execute(
        """
        SELECT
            pr.id,
            pr.user_id,
            pr.token_hash,
            pr.created_em,
            pr.expires_em,
            pr.used_em,
            u.email,
            u.nome,
            u.is_active
        FROM password_reset_tokens pr
        JOIN users u ON u.id = pr.user_id
        WHERE pr.token_hash = ?
        LIMIT 1
        """,
        (token_hash,),
    ).fetchone()
    if not row:
        return None
    if (row["used_em"] or "").strip():
        return None
    expires_dt = parse_iso_datetime(row["expires_em"] or "")
    if not expires_dt or expires_dt <= datetime.now():
        return None
    if not bool(row["is_active"]):
        return None
    return row


def get_password_cipher():
    global _password_cipher, _password_cipher_ready
    if _password_cipher_ready:
        return _password_cipher
    _password_cipher_ready = True
    if Fernet is None:
        return None

    key = ADMIN_PASSWORD_VIEW_KEY
    if not key:
        seed = (os.environ.get("FLASK_SECRET_KEY") or "csc-change-me").encode("utf-8")
        key = base64.urlsafe_b64encode(hashlib.sha256(seed).digest()).decode("utf-8")
    try:
        _password_cipher = Fernet(key.encode("utf-8"))
    except Exception:
        _password_cipher = None
    return _password_cipher


def build_password_view_cipher(raw_password: str) -> str:
    plain = (raw_password or "").strip()
    if not plain:
        return ""
    cipher = get_password_cipher()
    if cipher is None:
        return ""
    try:
        return cipher.encrypt(plain.encode("utf-8")).decode("utf-8")
    except Exception:
        return ""


def reveal_password_from_cipher(cipher_value: str) -> str:
    token = (cipher_value or "").strip()
    if not token:
        return ""
    cipher = get_password_cipher()
    if cipher is None:
        return ""
    try:
        return cipher.decrypt(token.encode("utf-8")).decode("utf-8")
    except InvalidToken:
        return ""
    except Exception:
        return ""


def get_master_admin_email() -> str:
    normalized = normalize_email(MASTER_ADMIN_EMAIL)
    if not normalized or "@" not in normalized:
        return "mangefestekaua@gmail.com"
    return normalized


def is_master_admin_email(email: str) -> bool:
    return normalize_email(email) == get_master_admin_email()


def row_email(row) -> str:
    if row is None:
        return ""
    if isinstance(row, dict):
        return str(row.get("email") or "")
    try:
        return str(row["email"] or "")
    except Exception:
        return ""


def is_master_admin_row(row) -> bool:
    return is_master_admin_email(row_email(row))


def password_hash_matches(password_hash: str, plain_password: str) -> bool:
    if not password_hash or not plain_password:
        return False
    try:
        return check_password_hash(password_hash, plain_password)
    except Exception:
        return False


def is_google_auth_enabled() -> bool:
    return bool(GOOGLE_CLIENT_ID and GOOGLE_CLIENT_SECRET)


def get_google_openid_config():
    global _google_openid_cache
    if _google_openid_cache is not None:
        return _google_openid_cache

    req = Request(
        GOOGLE_DISCOVERY_URL,
        headers={"User-Agent": "csc-insumos-auth/1.0", "Accept": "application/json"},
    )
    with urlopen(req, timeout=8) as response:
        payload = json.loads(response.read().decode("utf-8"))
    _google_openid_cache = payload
    return payload


def is_safe_next_url(next_url: str) -> bool:
    candidate = (next_url or "").strip()
    if not candidate:
        return False
    if not candidate.startswith("/") or candidate.startswith("//"):
        return False
    parsed = urlparse(candidate)
    if parsed.scheme or parsed.netloc:
        return False
    return True


def pick_next_url(*candidates: str) -> str:
    for candidate in candidates:
        if is_safe_next_url(candidate):
            return candidate.strip()
    return ""


def current_request_relative_url() -> str:
    path = request.path or "/"
    query = request.query_string.decode("utf-8", errors="ignore")
    return f"{path}?{query}" if query else path


def fetch_user_by_id(user_id):
    if not user_id:
        return None
    with closing(get_db()) as conn:
        row = conn.execute(
            """
            SELECT
                id,
                nome,
                email,
                provider,
                provider_sub,
                is_admin,
                is_active,
                created_em,
                updated_em,
                last_login_em,
                last_seen_em
            FROM users
            WHERE id = ?
            """,
            (user_id,),
        ).fetchone()
    if not row:
        return None
    user = dict(row)
    user["is_admin"] = bool(user.get("is_admin"))
    user["is_active"] = bool(user.get("is_active"))
    return user


def login_user(user_id: int):
    session["user_id"] = int(user_id)
    session["session_nonce"] = token_urlsafe(12)
    session["last_seen_write_ts"] = 0


def logout_user():
    session.pop("user_id", None)
    session.pop("session_nonce", None)
    session.pop("last_seen_write_ts", None)
    session.pop("oauth_google_state", None)
    session.pop("oauth_google_next", None)


def touch_user_activity(user):
    if not user:
        return
    user_id = int(user.get("id") or 0)
    if not user_id:
        return

    now_dt = datetime.now()
    now_ts = int(now_dt.timestamp())
    last_write_ts = int(session.get("last_seen_write_ts") or 0)
    if now_ts - last_write_ts < USER_LAST_SEEN_WRITE_INTERVAL_SECONDS:
        return

    ts = now_dt.isoformat()
    with closing(get_db()) as conn:
        conn.execute(
            "UPDATE users SET last_seen_em = ?, updated_em = ? WHERE id = ?",
            (ts, ts, user_id),
        )
        conn.commit()
    session["last_seen_write_ts"] = now_ts
    user["last_seen_em"] = ts


def resolve_post_login_target(user, fallback_endpoint: str = "index"):
    default_endpoint = "admin" if user and user.get("is_admin") else fallback_endpoint
    next_url = pick_next_url(
        request.form.get("next", ""),
        request.args.get("next", ""),
        session.pop("oauth_google_next", ""),
    )
    if next_url:
        return redirect(next_url)
    return redirect(url_for(default_endpoint))


def ensure_master_admin(conn, cur):
    normalized_email = get_master_admin_email()

    current = cur.execute(
        """
        SELECT id, nome, email, password_hash, password_view_cipher, is_admin, is_active
        FROM users
        WHERE lower(email) = lower(?)
        """,
        (normalized_email,),
    ).fetchone()

    ts = now_iso()
    if current:
        updates = []
        params = []
        if not current["nome"] and MASTER_ADMIN_NAME:
            updates.append("nome = ?")
            params.append(MASTER_ADMIN_NAME)
        if MASTER_ADMIN_PASSWORD and not password_hash_matches(
            current["password_hash"], MASTER_ADMIN_PASSWORD
        ):
            updates.append("password_hash = ?")
            params.append(generate_password_hash(MASTER_ADMIN_PASSWORD))
            updates.append("password_view_cipher = ?")
            params.append(build_password_view_cipher(MASTER_ADMIN_PASSWORD))
        elif MASTER_ADMIN_PASSWORD and not (current["password_view_cipher"] or "").strip():
            updates.append("password_view_cipher = ?")
            params.append(build_password_view_cipher(MASTER_ADMIN_PASSWORD))
        if not bool(current["is_admin"]):
            updates.append("is_admin = 1")
        if not bool(current["is_active"]):
            updates.append("is_active = 1")
        if (current["email"] or "").strip().lower() != normalized_email:
            updates.append("email = ?")
            params.append(normalized_email)
        updates.append("updated_em = ?")
        params.append(ts)
        if updates:
            params.append(current["id"])
            cur.execute(
                f"UPDATE users SET {', '.join(updates)} WHERE id = ?",
                params,
            )
        return

    password_hash = generate_password_hash(MASTER_ADMIN_PASSWORD or "Admin@123")
    cur.execute(
        """
        INSERT INTO users
            (nome, email, password_hash, password_view_cipher, provider, provider_sub, is_admin, is_active, created_em, updated_em, last_login_em, last_seen_em)
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """,
        (
            MASTER_ADMIN_NAME or "Administrador Master",
            normalized_email,
            password_hash,
            build_password_view_cipher(MASTER_ADMIN_PASSWORD),
            "local",
            "",
            1,
            1,
            ts,
            ts,
            None,
            None,
        ),
    )


def count_active_admins(cur) -> int:
    return cur.execute(
        "SELECT COUNT(1) FROM users WHERE is_admin = 1 AND is_active = 1"
    ).fetchone()[0]


def fetch_user_obra_ids(cur, user_id: int):
    if not user_id:
        return []
    rows = cur.execute(
        "SELECT obra_id FROM user_obras WHERE user_id = ? ORDER BY obra_id",
        (user_id,),
    ).fetchall()
    return [int(row["obra_id"]) for row in rows if row["obra_id"] is not None]


def filter_obras_for_user(cur, user, obras):
    if not user:
        return []
    if is_master_admin_email(user.get("email", "")):
        return obras

    user_id = int(user.get("id") or 0)
    allowed_ids = set(fetch_user_obra_ids(cur, user_id))
    if not allowed_ids:
        return []
    return [obra for obra in obras if int(obra.get("id") or 0) in allowed_ids]


def user_has_access_to_obra(cur, user, obra_id: int) -> bool:
    if not user or not obra_id:
        return False
    if is_master_admin_email(user.get("email", "")):
        return True
    user_id = int(user.get("id") or 0)
    return int(obra_id) in set(fetch_user_obra_ids(cur, user_id))


def init_db():
    with closing(get_db()) as conn:
        cur = conn.cursor()
        cur.execute(
            """
            CREATE TABLE IF NOT EXISTS obras (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                nome TEXT NOT NULL
            )
            """
        )
        cur.execute(
            """
            CREATE TABLE IF NOT EXISTS categorias (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                nome TEXT NOT NULL
            )
            """
        )
        cur.execute(
            """
            CREATE TABLE IF NOT EXISTS insumos (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                nome TEXT NOT NULL
            )
            """
        )
        cur.execute("PRAGMA table_info(insumos)")
        colunas = {row[1] for row in cur.fetchall()}
        if "categoria_id" not in colunas:
            cur.execute("ALTER TABLE insumos ADD COLUMN categoria_id INTEGER")
        if "foto_path" not in colunas:
            cur.execute("ALTER TABLE insumos ADD COLUMN foto_path TEXT")
        if "is_custom" not in colunas:
            cur.execute(
                "ALTER TABLE insumos ADD COLUMN is_custom INTEGER NOT NULL DEFAULT 0"
            )
        cur.execute(
            """
            CREATE TABLE IF NOT EXISTS unidades (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                nome TEXT NOT NULL
            )
            """
        )
        cur.execute(
            """
            CREATE TABLE IF NOT EXISTS insumo_unidades (
                insumo_id INTEGER NOT NULL,
                unidade_id INTEGER NOT NULL,
                PRIMARY KEY (insumo_id, unidade_id),
                FOREIGN KEY (insumo_id) REFERENCES insumos(id),
                FOREIGN KEY (unidade_id) REFERENCES unidades(id)
            )
            """
        )
        cur.execute(
            """
            CREATE TABLE IF NOT EXISTS especificacoes (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                insumo_id INTEGER NOT NULL,
                nome TEXT NOT NULL,
                FOREIGN KEY (insumo_id) REFERENCES insumos(id)
            )
            """
        )
        cur.execute(
            """
            CREATE TABLE IF NOT EXISTS requisicoes (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                obra_id INTEGER NOT NULL,
                criado_em TEXT NOT NULL,
                FOREIGN KEY (obra_id) REFERENCES obras(id)
            )
            """
        )
        cur.execute("PRAGMA table_info(requisicoes)")
        requisicao_colunas = {row[1] for row in cur.fetchall()}
        if "solicitante" not in requisicao_colunas:
            cur.execute("ALTER TABLE requisicoes ADD COLUMN solicitante TEXT")
        if "necessario_em" not in requisicao_colunas:
            cur.execute("ALTER TABLE requisicoes ADD COLUMN necessario_em TEXT")
        cur.execute(
            """
            CREATE TABLE IF NOT EXISTS requisicao_itens (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                requisicao_id INTEGER NOT NULL,
                insumo_id INTEGER NOT NULL,
                unidade TEXT NOT NULL,
                quantidade REAL NOT NULL,
                especificacao TEXT NOT NULL,
                apropriacao TEXT,
                necessario_em TEXT,
                FOREIGN KEY (requisicao_id) REFERENCES requisicoes(id),
                FOREIGN KEY (insumo_id) REFERENCES insumos(id)
            )
            """
        )
        cur.execute("PRAGMA table_info(requisicao_itens)")
        requisicao_item_colunas = {row[1] for row in cur.fetchall()}
        if "necessario_em" not in requisicao_item_colunas:
            cur.execute("ALTER TABLE requisicao_itens ADD COLUMN necessario_em TEXT")
        if "link_produto" not in requisicao_item_colunas:
            cur.execute("ALTER TABLE requisicao_itens ADD COLUMN link_produto TEXT")
        if "foto_path" not in requisicao_item_colunas:
            cur.execute("ALTER TABLE requisicao_itens ADD COLUMN foto_path TEXT")
        cur.execute(
            """
            CREATE TABLE IF NOT EXISTS apropriacoes (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                obra_id INTEGER NOT NULL,
                numero TEXT NOT NULL,
                nome TEXT NOT NULL DEFAULT '',
                FOREIGN KEY (obra_id) REFERENCES obras(id)
            )
            """
        )
        cur.execute("PRAGMA table_info(apropriacoes)")
        apropriacao_cols = {row[1] for row in cur.fetchall()}
        if "nome" not in apropriacao_cols:
            cur.execute("ALTER TABLE apropriacoes ADD COLUMN nome TEXT NOT NULL DEFAULT ''")

        cur.execute(
            """
            CREATE TABLE IF NOT EXISTS users (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                nome TEXT NOT NULL,
                email TEXT NOT NULL UNIQUE,
                password_hash TEXT,
                provider TEXT NOT NULL DEFAULT 'local',
                provider_sub TEXT,
                is_admin INTEGER NOT NULL DEFAULT 0,
                is_active INTEGER NOT NULL DEFAULT 1,
                created_em TEXT NOT NULL,
                updated_em TEXT NOT NULL,
                last_login_em TEXT,
                last_seen_em TEXT
            )
            """
        )
        cur.execute(
            """
            CREATE TABLE IF NOT EXISTS user_obras (
                user_id INTEGER NOT NULL,
                obra_id INTEGER NOT NULL,
                created_em TEXT NOT NULL,
                PRIMARY KEY (user_id, obra_id),
                FOREIGN KEY (user_id) REFERENCES users(id),
                FOREIGN KEY (obra_id) REFERENCES obras(id)
            )
            """
        )
        cur.execute(
            """
            CREATE TABLE IF NOT EXISTS password_reset_tokens (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                user_id INTEGER NOT NULL,
                token_hash TEXT NOT NULL UNIQUE,
                created_em TEXT NOT NULL,
                expires_em TEXT NOT NULL,
                used_em TEXT,
                requested_ip TEXT,
                requested_ua TEXT,
                FOREIGN KEY (user_id) REFERENCES users(id)
            )
            """
        )
        cur.execute("PRAGMA table_info(users)")
        user_cols = {row[1] for row in cur.fetchall()}
        if "nome" not in user_cols:
            cur.execute("ALTER TABLE users ADD COLUMN nome TEXT NOT NULL DEFAULT ''")
        if "email" not in user_cols:
            cur.execute("ALTER TABLE users ADD COLUMN email TEXT NOT NULL DEFAULT ''")
        if "password_hash" not in user_cols:
            cur.execute("ALTER TABLE users ADD COLUMN password_hash TEXT")
        if "password_view_cipher" not in user_cols:
            cur.execute("ALTER TABLE users ADD COLUMN password_view_cipher TEXT")
        if "provider" not in user_cols:
            cur.execute("ALTER TABLE users ADD COLUMN provider TEXT NOT NULL DEFAULT 'local'")
        if "provider_sub" not in user_cols:
            cur.execute("ALTER TABLE users ADD COLUMN provider_sub TEXT")
        if "is_admin" not in user_cols:
            cur.execute("ALTER TABLE users ADD COLUMN is_admin INTEGER NOT NULL DEFAULT 0")
        if "is_active" not in user_cols:
            cur.execute("ALTER TABLE users ADD COLUMN is_active INTEGER NOT NULL DEFAULT 1")
        if "created_em" not in user_cols:
            cur.execute("ALTER TABLE users ADD COLUMN created_em TEXT NOT NULL DEFAULT ''")
        if "updated_em" not in user_cols:
            cur.execute("ALTER TABLE users ADD COLUMN updated_em TEXT NOT NULL DEFAULT ''")
        if "last_login_em" not in user_cols:
            cur.execute("ALTER TABLE users ADD COLUMN last_login_em TEXT")
        if "last_seen_em" not in user_cols:
            cur.execute("ALTER TABLE users ADD COLUMN last_seen_em TEXT")
        cur.execute("CREATE UNIQUE INDEX IF NOT EXISTS idx_users_email ON users(email)")
        cur.execute(
            "CREATE UNIQUE INDEX IF NOT EXISTS idx_user_obras_pair ON user_obras(user_id, obra_id)"
        )
        cur.execute("CREATE INDEX IF NOT EXISTS idx_user_obras_user ON user_obras(user_id)")
        cur.execute("CREATE INDEX IF NOT EXISTS idx_user_obras_obra ON user_obras(obra_id)")
        cur.execute(
            "CREATE UNIQUE INDEX IF NOT EXISTS idx_password_reset_token_hash ON password_reset_tokens(token_hash)"
        )
        cur.execute(
            "CREATE INDEX IF NOT EXISTS idx_password_reset_user ON password_reset_tokens(user_id)"
        )
        cur.execute(
            """
            DELETE FROM user_obras
            WHERE user_id NOT IN (SELECT id FROM users)
               OR obra_id NOT IN (SELECT id FROM obras)
            """
        )
        cleanup_cutoff = (datetime.now() - timedelta(days=15)).isoformat()
        cur.execute(
            """
            DELETE FROM password_reset_tokens
            WHERE user_id NOT IN (SELECT id FROM users)
               OR COALESCE(TRIM(token_hash), '') = ''
               OR COALESCE(TRIM(expires_em), '') = ''
               OR (used_em IS NOT NULL AND used_em < ?)
               OR expires_em < ?
            """,
            (cleanup_cutoff, cleanup_cutoff),
        )
        cur.execute(
            """
            CREATE UNIQUE INDEX IF NOT EXISTS idx_users_provider_sub
            ON users(provider, provider_sub)
            WHERE provider_sub IS NOT NULL AND TRIM(provider_sub) <> ''
            """
        )
        cur.execute("UPDATE users SET email = lower(trim(email)) WHERE email IS NOT NULL")
        timestamp = now_iso()
        cur.execute(
            """
            UPDATE users
            SET
                provider = COALESCE(NULLIF(provider, ''), 'local'),
                created_em = CASE WHEN COALESCE(TRIM(created_em), '') = '' THEN ? ELSE created_em END,
                updated_em = CASE WHEN COALESCE(TRIM(updated_em), '') = '' THEN ? ELSE updated_em END,
                nome = CASE WHEN COALESCE(TRIM(nome), '') = '' THEN email ELSE nome END,
                last_seen_em = CASE WHEN COALESCE(TRIM(last_seen_em), '') = '' THEN last_login_em ELSE last_seen_em END
            """,
            (timestamp, timestamp),
        )
        ensure_master_admin(conn, cur)
        conn.commit()


def rows_to_dicts(rows):
    return [dict(row) for row in rows]


def is_valid_apropriacao(valor: str) -> bool:
    if not valor:
        return False
    return re.fullmatch(r"\d+(?:\.\d+)*", valor) is not None


def normalize_text(valor: str) -> str:
    texto = (valor or "").strip().lower()
    texto = unicodedata.normalize("NFKD", texto)
    texto = "".join(ch for ch in texto if not unicodedata.combining(ch))
    return texto


def resolve_header_index(headers):
    aliases = {
        "insumo": {"insumo", "material", "item"},
        "descricao": {"descricao", "especificacao", "especificacao tecnica", "desc"},
    }
    normalized = [normalize_text(h) for h in headers]

    idx = {}
    for field, opts in aliases.items():
        for i, col in enumerate(normalized):
            if col in opts:
                idx[field] = i
                break
    if "insumo" not in idx or "descricao" not in idx:
        return None
    return idx


def parse_specs_upload(file_storage):
    filename = (file_storage.filename or "").strip()
    if not filename:
        return None, "Selecione um arquivo CSV ou XLSX."

    ext = os.path.splitext(filename.lower())[1]
    rows = []
    try:
        if ext == ".csv":
            raw = file_storage.read()
            text = raw.decode("utf-8-sig", errors="replace")
            lines = text.splitlines()
            if not lines:
                return None, "CSV vazio."
            header_line = lines[0]
            delimiters = [",", ";", "\t"]
            delimiter = max(delimiters, key=lambda item: header_line.count(item))
            if header_line.count(delimiter) == 0:
                delimiter = ","
            reader = csv.reader(lines, delimiter=delimiter)
            headers = next(reader, None)
            if not headers:
                return None, "CSV sem cabecalho."
            idx = resolve_header_index(headers)
            if not idx:
                return None, "Colunas esperadas: insumo e descricao/especificacao."
            for row in reader:
                if len(row) <= max(idx.values()):
                    continue
                insumo = row[idx["insumo"]].strip()
                descricao = row[idx["descricao"]].strip()
                if insumo and descricao:
                    rows.append((insumo, descricao))
        elif ext == ".xlsx":
            if openpyxl is None:
                return None, "Instale openpyxl para importar XLSX."
            wb = openpyxl.load_workbook(file_storage, read_only=True, data_only=True)
            ws = wb.active
            all_rows = list(ws.iter_rows(values_only=True))
            if not all_rows:
                return None, "XLSX vazio."
            headers = [str(v or "") for v in all_rows[0]]
            idx = resolve_header_index(headers)
            if not idx:
                return None, "Colunas esperadas: insumo e descricao/especificacao."
            for r in all_rows[1:]:
                if len(r) <= max(idx.values()):
                    continue
                insumo = str(r[idx["insumo"]] or "").strip()
                descricao = str(r[idx["descricao"]] or "").strip()
                if insumo and descricao:
                    rows.append((insumo, descricao))
        else:
            return None, "Formato invalido. Use CSV ou XLSX."
    except Exception:
        return None, "Nao foi possivel ler o arquivo."

    if not rows:
        return None, "Nenhuma linha valida encontrada."
    return rows, None


def split_multi_values(valor: str):
    return [item.strip() for item in re.split(r"[,\n;|]+", valor or "") if item.strip()]


def parse_apropriacao_entry(raw: str):
    texto = (raw or "").strip()
    if not texto:
        return "", ""

    texto = texto.replace(",", ".")

    for sep in ("|", ";", "\t", " - ", " -", "- "):
        if sep not in texto:
            continue
        left, right = texto.split(sep, 1)
        left = left.strip()
        right = right.strip()
        if is_valid_apropriacao(left):
            return left, right
        if is_valid_apropriacao(right):
            return right, left
        return left, right

    codigo_inicio = re.match(r"^(\d+(?:\.\d+)*)(?:\s+(.+))?$", texto)
    if codigo_inicio:
        codigo = (codigo_inicio.group(1) or "").strip()
        nome = (codigo_inicio.group(2) or "").strip()
        return codigo, nome

    codigo_fim = re.match(r"^(.+?)\s+(\d+(?:\.\d+)*)$", texto)
    if codigo_fim:
        nome = (codigo_fim.group(1) or "").strip()
        codigo = (codigo_fim.group(2) or "").strip()
        return codigo, nome

    return texto, ""


def build_apropriacao_nome_lookup(obra_id) -> dict:
    if not obra_id:
        return {}
    with closing(get_db()) as conn:
        rows = conn.execute(
            """
            SELECT numero, TRIM(COALESCE(nome, '')) AS nome
            FROM apropriacoes
            WHERE obra_id = ?
            """,
            (obra_id,),
        ).fetchall()

    lookup = {}
    ambiguos = set()
    for row in rows:
        codigo = (row["numero"] or "").strip().replace(",", ".")
        nome = (row["nome"] or "").strip()
        if not codigo or not nome:
            continue
        chave = normalize_text(nome)
        if not chave:
            continue
        atual = lookup.get(chave)
        if atual and atual != codigo:
            ambiguos.add(chave)
            continue
        lookup[chave] = codigo

    for chave in ambiguos:
        lookup.pop(chave, None)
    return lookup


def resolve_apropriacao_codigo(raw_value: str, nome_lookup: dict) -> str:
    texto = (raw_value or "").strip()
    if not texto:
        return ""

    codigo, nome = parse_apropriacao_entry(texto)
    if codigo and is_valid_apropriacao(codigo):
        return codigo

    chave_direta = normalize_text(texto)
    if chave_direta and chave_direta in nome_lookup:
        return nome_lookup[chave_direta]

    chave_nome = normalize_text(nome)
    if chave_nome and chave_nome in nome_lookup:
        return nome_lookup[chave_nome]

    return ""


def resolve_insumos_upload_index(headers):
    normalized = [normalize_text(h) for h in headers]

    def pick(aliases):
        for idx, col in enumerate(normalized):
            if col in aliases:
                return idx
        return None

    idx_insumo = pick({"insumo", "nome", "material", "item"})
    if idx_insumo is None:
        return None
    return {
        "insumo": idx_insumo,
        "categoria": pick({"categoria", "grupo"}),
        "unidades": pick({"unidades", "unidade", "metragem", "medida"}),
        "especificacoes": pick(
            {
                "especificacoes",
                "especificacao",
                "descricao",
                "descricoes",
                "desc",
            }
        ),
    }


def parse_insumos_upload(file_storage):
    filename = (file_storage.filename or "").strip()
    if not filename:
        return None, "Selecione um arquivo CSV ou XLSX."

    ext = os.path.splitext(filename.lower())[1]
    rows = []
    try:
        if ext == ".csv":
            raw = file_storage.read()
            text = raw.decode("utf-8-sig", errors="replace")
            lines = text.splitlines()
            if not lines:
                return None, "CSV vazio."
            header_line = lines[0]
            delimiters = [",", ";", "\t"]
            delimiter = max(delimiters, key=lambda item: header_line.count(item))
            if header_line.count(delimiter) == 0:
                delimiter = ","
            reader = csv.reader(lines, delimiter=delimiter)
            headers = next(reader, None)
            if not headers:
                return None, "CSV sem cabecalho."
            idx = resolve_insumos_upload_index(headers)
            if not idx:
                return None, "Coluna obrigatoria nao encontrada: nome/insumo."
            max_idx = max(v for v in idx.values() if v is not None)
            for row in reader:
                if len(row) <= max_idx:
                    continue
                nome = str(row[idx["insumo"]] or "").strip()
                if not nome:
                    continue
                categoria = (
                    str(row[idx["categoria"]] or "").strip()
                    if idx["categoria"] is not None and len(row) > idx["categoria"]
                    else ""
                )
                unidades_raw = (
                    str(row[idx["unidades"]] or "").strip()
                    if idx["unidades"] is not None and len(row) > idx["unidades"]
                    else ""
                )
                especificacoes_raw = (
                    str(row[idx["especificacoes"]] or "").strip()
                    if idx["especificacoes"] is not None
                    and len(row) > idx["especificacoes"]
                    else ""
                )
                rows.append(
                    {
                        "nome": nome,
                        "categoria": categoria,
                        "unidades": split_multi_values(unidades_raw),
                        "especificacoes": split_multi_values(especificacoes_raw),
                    }
                )
        elif ext == ".xlsx":
            if openpyxl is None:
                return None, "Instale openpyxl para importar XLSX."
            wb = openpyxl.load_workbook(file_storage, read_only=True, data_only=True)
            ws = wb.active
            for sheet in wb.worksheets:
                if normalize_text(sheet.title) in {
                    "insumos",
                    "insumo",
                    "materiais",
                    "material",
                }:
                    ws = sheet
                    break

            all_rows = list(ws.iter_rows(values_only=True))
            if not all_rows:
                return None, "XLSX vazio."
            headers = [str(v or "") for v in all_rows[0]]
            idx = resolve_insumos_upload_index(headers)
            if not idx:
                return None, "Coluna obrigatoria nao encontrada: nome/insumo."
            for r in all_rows[1:]:
                if len(r) <= idx["insumo"]:
                    continue
                nome = str(r[idx["insumo"]] or "").strip()
                if not nome:
                    continue
                categoria = (
                    str(r[idx["categoria"]] or "").strip()
                    if idx["categoria"] is not None and len(r) > idx["categoria"]
                    else ""
                )
                unidades_raw = (
                    str(r[idx["unidades"]] or "").strip()
                    if idx["unidades"] is not None and len(r) > idx["unidades"]
                    else ""
                )
                especificacoes_raw = (
                    str(r[idx["especificacoes"]] or "").strip()
                    if idx["especificacoes"] is not None
                    and len(r) > idx["especificacoes"]
                    else ""
                )
                rows.append(
                    {
                        "nome": nome,
                        "categoria": categoria,
                        "unidades": split_multi_values(unidades_raw),
                        "especificacoes": split_multi_values(especificacoes_raw),
                    }
                )
        else:
            return None, "Formato invalido. Use CSV ou XLSX."
    except Exception:
        return None, "Nao foi possivel ler o arquivo."

    if not rows:
        return None, "Nenhuma linha valida encontrada."

    payload = {"insumos": [], "especificacoes": []}
    for row in rows:
        payload["insumos"].append(
            {
                "nome": row["nome"],
                "categoria": row["categoria"],
                "unidades": row["unidades"],
            }
        )
        for descricao in row["especificacoes"]:
            payload["especificacoes"].append(
                {"insumo": row["nome"], "descricao": descricao}
            )
    return payload, None


def parse_master_upload(file_storage):
    filename = (file_storage.filename or "").strip()
    if not filename:
        return None, "Selecione um arquivo CSV ou XLSX."

    ext = os.path.splitext(filename.lower())[1]
    payload = {
        "obras": [],
        "categorias": [],
        "unidades": [],
        "insumos": [],
        "especificacoes": [],
        "apropriacoes": [],
    }

    def add_nome(lista_nome, valor):
        nome = str(valor or "").strip()
        if nome:
            lista_nome.append(nome)

    def pick_idx(headers_norm, aliases):
        for idx, col in enumerate(headers_norm):
            if col in aliases:
                return idx
        return None

    def parse_sheet(titulo, rows):
        if not rows:
            return None
        headers = [str(v or "") for v in rows[0]]
        headers_norm = [normalize_text(h) for h in headers]
        title_norm = normalize_text(titulo)

        if title_norm in {"obras", "obra"}:
            idx_nome = pick_idx(headers_norm, {"nome", "obra", "projeto"})
            if idx_nome is None:
                idx_nome = 0
            for row in rows[1:]:
                if len(row) > idx_nome:
                    add_nome(payload["obras"], row[idx_nome])
            return None

        if title_norm in {"categorias", "categoria"}:
            idx_nome = pick_idx(headers_norm, {"nome", "categoria", "grupo"})
            if idx_nome is None:
                idx_nome = 0
            for row in rows[1:]:
                if len(row) > idx_nome:
                    add_nome(payload["categorias"], row[idx_nome])
            return None

        if title_norm in {"unidades", "unidade", "metragem", "medidas", "medida"}:
            idx_nome = pick_idx(headers_norm, {"nome", "unidade", "metragem", "medida"})
            if idx_nome is None:
                idx_nome = 0
            for row in rows[1:]:
                if len(row) > idx_nome:
                    add_nome(payload["unidades"], row[idx_nome])
            return None

        if title_norm in {"insumos", "insumo", "materiais", "material"}:
            idx_nome = pick_idx(headers_norm, {"insumo", "nome", "material"})
            idx_categoria = pick_idx(headers_norm, {"categoria", "grupo"})
            idx_unidades = pick_idx(headers_norm, {"unidades", "unidade", "metragem", "medida"})
            if idx_nome is None:
                return "Aba insumos sem coluna de nome/insumo."
            for row in rows[1:]:
                if len(row) <= idx_nome:
                    continue
                nome = str(row[idx_nome] or "").strip()
                if not nome:
                    continue
                categoria = ""
                if idx_categoria is not None and len(row) > idx_categoria:
                    categoria = str(row[idx_categoria] or "").strip()
                unidades_raw = ""
                if idx_unidades is not None and len(row) > idx_unidades:
                    unidades_raw = str(row[idx_unidades] or "").strip()
                payload["insumos"].append(
                    {
                        "nome": nome,
                        "categoria": categoria,
                        "unidades": split_multi_values(unidades_raw),
                    }
                )
            return None

        if title_norm in {"especificacoes", "especificacao", "descricoes", "descricao"}:
            idx_insumo = pick_idx(headers_norm, {"insumo", "nome", "material"})
            idx_desc = pick_idx(headers_norm, {"descricao", "especificacao", "desc"})
            if idx_insumo is None or idx_desc is None:
                return "Aba especificacoes precisa de colunas insumo e descricao."
            for row in rows[1:]:
                if len(row) <= max(idx_insumo, idx_desc):
                    continue
                insumo = str(row[idx_insumo] or "").strip()
                descricao = str(row[idx_desc] or "").strip()
                if insumo and descricao:
                    payload["especificacoes"].append(
                        {"insumo": insumo, "descricao": descricao}
                    )
            return None

        if title_norm in {"apropriacoes", "apropriacao"}:
            idx_obra = pick_idx(headers_norm, {"obra", "projeto"})
            idx_numero = pick_idx(headers_norm, {"numero", "apropriacao", "aprop", "codigo"})
            idx_nome = pick_idx(headers_norm, {"nome", "descricao", "titulo"})
            if idx_obra is None or idx_numero is None:
                return "Aba apropriacoes precisa de colunas obra e numero."
            for row in rows[1:]:
                if len(row) <= max(idx_obra, idx_numero):
                    continue
                obra = str(row[idx_obra] or "").strip()
                numero = str(row[idx_numero] or "").strip()
                nome = ""
                if idx_nome is not None and len(row) > idx_nome:
                    nome = str(row[idx_nome] or "").strip()
                if obra and numero:
                    payload["apropriacoes"].append(
                        {"obra": obra, "numero": numero, "nome": nome}
                    )
            return None

        return None

    try:
        if ext == ".xlsx":
            if openpyxl is None:
                return None, "Instale openpyxl para importar XLSX."
            wb = openpyxl.load_workbook(file_storage, read_only=True, data_only=True)
            for ws in wb.worksheets:
                rows = list(ws.iter_rows(values_only=True))
                erro = parse_sheet(ws.title, rows)
                if erro:
                    return None, erro
        elif ext == ".csv":
            raw = file_storage.read()
            text = raw.decode("utf-8-sig", errors="replace")
            reader = csv.DictReader(text.splitlines())
            if not reader.fieldnames:
                return None, "CSV sem cabecalho."
            field_map = {normalize_text(c): c for c in reader.fieldnames}
            if "tipo" not in field_map:
                return (
                    None,
                    "CSV deve ter coluna tipo (obra, categoria, unidade, insumo, especificacao, apropriacao).",
                )
            tipo_col = field_map["tipo"]
            for row in reader:
                tipo = normalize_text(row.get(tipo_col, ""))
                if tipo in {"obra", "obras"}:
                    add_nome(payload["obras"], row.get(field_map.get("nome", ""), ""))
                elif tipo in {"categoria", "categorias"}:
                    add_nome(payload["categorias"], row.get(field_map.get("nome", ""), ""))
                elif tipo in {"unidade", "unidades", "metragem"}:
                    add_nome(payload["unidades"], row.get(field_map.get("nome", ""), ""))
                elif tipo in {"insumo", "insumos", "material", "materiais"}:
                    nome = (row.get(field_map.get("nome", ""), "") or "").strip()
                    categoria = (row.get(field_map.get("categoria", ""), "") or "").strip()
                    unidades = split_multi_values(
                        row.get(field_map.get("unidades", ""), "")
                        or row.get(field_map.get("unidade", ""), "")
                    )
                    if nome:
                        payload["insumos"].append(
                            {
                                "nome": nome,
                                "categoria": categoria,
                                "unidades": unidades,
                            }
                        )
                elif tipo in {"especificacao", "especificacoes", "descricao", "descricoes"}:
                    insumo = (row.get(field_map.get("insumo", ""), "") or "").strip()
                    descricao = (
                        row.get(field_map.get("descricao", ""), "")
                        or row.get(field_map.get("especificacao", ""), "")
                        or ""
                    ).strip()
                    if insumo and descricao:
                        payload["especificacoes"].append(
                            {"insumo": insumo, "descricao": descricao}
                        )
                elif tipo in {"apropriacao", "apropriacoes"}:
                    obra = (row.get(field_map.get("obra", ""), "") or "").strip()
                    numero = (
                        row.get(field_map.get("numero", ""), "")
                        or row.get(field_map.get("apropriacao", ""), "")
                        or ""
                    ).strip()
                    nome = (
                        row.get(field_map.get("nome", ""), "")
                        or row.get(field_map.get("descricao", ""), "")
                        or ""
                    ).strip()
                    if obra and numero:
                        payload["apropriacoes"].append(
                            {"obra": obra, "numero": numero, "nome": nome}
                        )
        else:
            return None, "Formato invalido. Use CSV ou XLSX."
    except Exception:
        return None, "Nao foi possivel ler o arquivo."

    total = (
        len(payload["obras"])
        + len(payload["categorias"])
        + len(payload["unidades"])
        + len(payload["insumos"])
        + len(payload["especificacoes"])
        + len(payload["apropriacoes"])
    )
    if total == 0:
        return None, "Nenhuma linha valida encontrada no arquivo."
    return payload, None


def import_master_payload(payload):
    stats = {
        "obras_novas": 0,
        "categorias_novas": 0,
        "unidades_novas": 0,
        "insumos_novos": 0,
        "insumos_atualizados": 0,
        "vinculos_unidade_novos": 0,
        "especificacoes_novas": 0,
        "apropriacoes_novas": 0,
        "linhas_invalidas": 0,
    }

    with closing(get_db()) as conn:
        cur = conn.cursor()

        obra_map = {
            normalize_text(row["nome"]): row["id"]
            for row in cur.execute("SELECT id, nome FROM obras").fetchall()
        }
        categoria_map = {
            normalize_text(row["nome"]): row["id"]
            for row in cur.execute("SELECT id, nome FROM categorias").fetchall()
        }
        unidade_map = {
            normalize_text(row["nome"]): row["id"]
            for row in cur.execute("SELECT id, nome FROM unidades").fetchall()
        }
        insumo_rows = cur.execute(
            "SELECT id, nome, categoria_id FROM insumos WHERE COALESCE(is_custom, 0) = 0"
        ).fetchall()
        insumo_map = {normalize_text(row["nome"]): row["id"] for row in insumo_rows}
        insumo_categoria_atual = {
            row["id"]: row["categoria_id"] for row in insumo_rows
        }

        vinculos_existentes = {
            (row["insumo_id"], row["unidade_id"])
            for row in cur.execute(
                "SELECT insumo_id, unidade_id FROM insumo_unidades"
            ).fetchall()
        }
        specs_existentes = {}
        for row in cur.execute("SELECT insumo_id, nome FROM especificacoes").fetchall():
            specs_existentes.setdefault(row["insumo_id"], set()).add(
                normalize_text(row["nome"])
            )
        aprop_existentes = {}
        for row in cur.execute(
            "SELECT obra_id, numero, COALESCE(nome, '') AS nome FROM apropriacoes"
        ).fetchall():
            aprop_existentes.setdefault(row["obra_id"], {})[row["numero"]] = (
                row["nome"] or ""
            ).strip()

        def ensure_obra(nome):
            key = normalize_text(nome)
            if not key:
                return None
            obra_id = obra_map.get(key)
            if obra_id is None:
                cur.execute("INSERT INTO obras (nome) VALUES (?)", (nome,))
                obra_id = cur.lastrowid
                obra_map[key] = obra_id
                stats["obras_novas"] += 1
            aprop_existentes.setdefault(obra_id, {})
            return obra_id

        def ensure_categoria(nome):
            key = normalize_text(nome)
            if not key:
                return None
            categoria_id = categoria_map.get(key)
            if categoria_id is None:
                cur.execute("INSERT INTO categorias (nome) VALUES (?)", (nome,))
                categoria_id = cur.lastrowid
                categoria_map[key] = categoria_id
                stats["categorias_novas"] += 1
            return categoria_id

        def ensure_unidade(nome):
            key = normalize_text(nome)
            if not key:
                return None
            unidade_id = unidade_map.get(key)
            if unidade_id is None:
                cur.execute("INSERT INTO unidades (nome) VALUES (?)", (nome,))
                unidade_id = cur.lastrowid
                unidade_map[key] = unidade_id
                stats["unidades_novas"] += 1
            return unidade_id

        def ensure_insumo(nome, categoria_id=None):
            key = normalize_text(nome)
            if not key:
                return None
            insumo_id = insumo_map.get(key)
            if insumo_id is None:
                cur.execute(
                    "INSERT INTO insumos (nome, categoria_id) VALUES (?, ?)",
                    (nome, categoria_id),
                )
                insumo_id = cur.lastrowid
                insumo_map[key] = insumo_id
                insumo_categoria_atual[insumo_id] = categoria_id
                stats["insumos_novos"] += 1
            elif categoria_id and insumo_categoria_atual.get(insumo_id) != categoria_id:
                cur.execute(
                    "UPDATE insumos SET categoria_id = ? WHERE id = ?",
                    (categoria_id, insumo_id),
                )
                insumo_categoria_atual[insumo_id] = categoria_id
                stats["insumos_atualizados"] += 1
            return insumo_id

        for nome in payload.get("obras", []):
            if not ensure_obra(nome):
                stats["linhas_invalidas"] += 1

        for nome in payload.get("categorias", []):
            if not ensure_categoria(nome):
                stats["linhas_invalidas"] += 1

        for nome in payload.get("unidades", []):
            if not ensure_unidade(nome):
                stats["linhas_invalidas"] += 1

        for item in payload.get("insumos", []):
            nome = (item.get("nome") or "").strip()
            if not nome:
                stats["linhas_invalidas"] += 1
                continue
            categoria_nome = (item.get("categoria") or "").strip()
            categoria_id = ensure_categoria(categoria_nome) if categoria_nome else None
            insumo_id = ensure_insumo(nome, categoria_id)
            if not insumo_id:
                stats["linhas_invalidas"] += 1
                continue
            for unidade_nome in item.get("unidades") or []:
                unidade_id = ensure_unidade(unidade_nome)
                if not unidade_id:
                    continue
                par = (insumo_id, unidade_id)
                if par in vinculos_existentes:
                    continue
                cur.execute(
                    """
                    INSERT OR IGNORE INTO insumo_unidades (insumo_id, unidade_id)
                    VALUES (?, ?)
                    """,
                    par,
                )
                vinculos_existentes.add(par)
                stats["vinculos_unidade_novos"] += 1

        for spec in payload.get("especificacoes", []):
            insumo_nome = (spec.get("insumo") or "").strip()
            descricao = (spec.get("descricao") or "").strip()
            if not insumo_nome or not descricao:
                stats["linhas_invalidas"] += 1
                continue
            insumo_id = ensure_insumo(insumo_nome)
            if not insumo_id:
                stats["linhas_invalidas"] += 1
                continue
            desc_key = normalize_text(descricao)
            existentes = specs_existentes.setdefault(insumo_id, set())
            if desc_key in existentes:
                continue
            cur.execute(
                "INSERT INTO especificacoes (insumo_id, nome) VALUES (?, ?)",
                (insumo_id, descricao),
            )
            existentes.add(desc_key)
            stats["especificacoes_novas"] += 1

        for aprop in payload.get("apropriacoes", []):
            obra_nome = (aprop.get("obra") or "").strip()
            numero = (aprop.get("numero") or "").strip().replace(",", ".")
            nome_aprop = (aprop.get("nome") or "").strip()
            if not obra_nome or not numero:
                stats["linhas_invalidas"] += 1
                continue
            if not is_valid_apropriacao(numero):
                stats["linhas_invalidas"] += 1
                continue
            obra_id = ensure_obra(obra_nome)
            if not obra_id:
                stats["linhas_invalidas"] += 1
                continue
            existentes = aprop_existentes.setdefault(obra_id, {})
            nome_existente = existentes.get(numero)
            if nome_existente is not None:
                if nome_aprop and normalize_text(nome_existente) != normalize_text(nome_aprop):
                    cur.execute(
                        """
                        UPDATE apropriacoes
                        SET nome = ?
                        WHERE obra_id = ? AND numero = ?
                        """,
                        (nome_aprop, obra_id, numero),
                    )
                    existentes[numero] = nome_aprop
                continue
            cur.execute(
                "INSERT INTO apropriacoes (obra_id, numero, nome) VALUES (?, ?, ?)",
                (obra_id, numero, nome_aprop),
            )
            existentes[numero] = nome_aprop
            stats["apropriacoes_novas"] += 1

        conn.commit()
    return stats


def redirect_next(default_endpoint: str):
    next_url = pick_next_url(request.form.get("next", ""), request.args.get("next", ""))
    if next_url:
        return redirect(next_url)
    return redirect(url_for(default_endpoint))


def format_date_for_sheet(valor: str, fallback: str = "") -> str:
    for raw in (valor, fallback):
        if not raw:
            continue
        try:
            dt = datetime.fromisoformat(raw)
            return dt.strftime("%d/%m/%Y")
        except ValueError:
            pass
        try:
            dt = datetime.strptime(raw, "%Y-%m-%d")
            return dt.strftime("%d/%m/%Y")
        except ValueError:
            pass
    return ""


def parse_iso_date(valor: str):
    raw = (valor or "").strip()
    if not raw:
        return None
    try:
        return datetime.strptime(raw, "%Y-%m-%d").date()
    except ValueError:
        return None


def get_lead_days_for_item_count(total_itens: int) -> int:
    for max_items, days in REQUEST_LEAD_RULES:
        if total_itens <= max_items:
            return days
    return REQUEST_LEAD_FALLBACK_DAYS


def build_necessario_min_date(total_itens: int, base_date=None):
    ref_date = base_date or datetime.now().date()
    return ref_date + timedelta(days=get_lead_days_for_item_count(total_itens))


def parse_int_list(values):
    parsed = []
    for value in values:
        try:
            parsed.append(int(value))
        except (TypeError, ValueError):
            continue
    return sorted(set(parsed))


def get_or_create_unidades(cur, nomes):
    linhas = [nome.strip() for nome in nomes if nome and nome.strip()]
    if not linhas:
        return []

    existentes_rows = cur.execute("SELECT id, nome FROM unidades").fetchall()
    existentes = {normalize_text(row["nome"]): row["id"] for row in existentes_rows}
    ids = []
    for nome in linhas:
        key = normalize_text(nome)
        unidade_id = existentes.get(key)
        if unidade_id is None:
            cur.execute("INSERT INTO unidades (nome) VALUES (?)", (nome,))
            unidade_id = cur.lastrowid
            existentes[key] = unidade_id
        ids.append(unidade_id)
    return sorted(set(ids))


def is_valid_image_rel_path(rel_path: str) -> bool:
    if not rel_path:
        return False
    normalized = rel_path.replace("\\", "/")
    return normalized.startswith("uploads/insumos/") or normalized.startswith(
        "uploads/requisicoes/"
    )


def remove_image_file(rel_path: str):
    if not is_valid_image_rel_path(rel_path):
        return
    normalized = rel_path.replace("\\", "/")
    abs_path = os.path.join(app.root_path, "static", normalized.replace("/", os.sep))
    if os.path.isfile(abs_path):
        os.remove(abs_path)


def remove_insumo_image(rel_path: str):
    remove_image_file(rel_path)


def resolve_uploaded_image_path(path_value: str) -> str:
    candidate = (path_value or "").strip()
    if not candidate:
        return ""

    if os.path.isabs(candidate) and os.path.isfile(candidate):
        return candidate

    normalized = candidate.replace("\\", "/").lstrip("/")
    normalized_lower = normalized.lower()
    if normalized_lower.startswith("static/"):
        normalized = normalized.split("/", 1)[1]

    possible_paths = [
        os.path.join(app.root_path, "static", normalized.replace("/", os.sep)),
        os.path.join(app.root_path, normalized.replace("/", os.sep)),
    ]
    for abs_path in possible_paths:
        if os.path.isfile(abs_path):
            return abs_path
    return ""


def parse_image_payload(file_storage, clipboard_data: str, allowed_extensions):
    image_bytes = None
    ext = ""

    if file_storage and (file_storage.filename or "").strip():
        filename = file_storage.filename.strip()
        ext = os.path.splitext(filename.lower())[1]
        if ext not in allowed_extensions:
            return None, "", "Formato de imagem invalido. Use PNG, JPG, JPEG, JFIF ou WEBP."
        image_bytes = file_storage.read()
    else:
        data = (clipboard_data or "").strip()
        if data:
            match = re.match(r"^data:image/([a-zA-Z0-9.+-]+);base64,(.+)$", data)
            if not match:
                return None, "", "Nao foi possivel ler a imagem colada."
            subtype = match.group(1).lower()
            ext_map = {
                "png": ".png",
                "jpeg": ".jpg",
                "jpg": ".jpg",
                "pjpeg": ".jpg",
                "jfif": ".jpg",
                "webp": ".webp",
                "x-png": ".png",
            }
            ext = ext_map.get(subtype)
            if not ext:
                return None, "", "Tipo de imagem nao suportado."
            try:
                image_bytes = base64.b64decode(match.group(2), validate=True)
            except (binascii.Error, ValueError):
                return None, "", "Imagem colada invalida."

    if not image_bytes:
        return None, "", ""
    if len(image_bytes) > MAX_IMAGE_SIZE_BYTES:
        return None, "", "Imagem muito grande. Limite de 8MB."
    return image_bytes, ext, ""


def polish_image_bytes(image_bytes: bytes, ext: str, logo: bool = False):
    if Image is None:
        return image_bytes, ext, ""

    try:
        with Image.open(BytesIO(image_bytes)) as img:
            if ImageOps is not None:
                img = ImageOps.exif_transpose(img)

            if hasattr(Image, "Resampling"):
                resample = Image.Resampling.LANCZOS
            else:
                resample = getattr(Image, "LANCZOS", Image.BICUBIC)

            if logo:
                max_logo_size = (620, 240)
                if img.width > max_logo_size[0] or img.height > max_logo_size[1]:
                    img.thumbnail(max_logo_size, resample)

                output = BytesIO()
                if img.mode in {"RGBA", "LA", "P"}:
                    img = img.convert("RGBA")
                else:
                    img = img.convert("RGB")
                img.save(output, format="PNG", optimize=True)
                return output.getvalue(), ".png", ""

            max_side = REQUISICAO_IMAGE_MAX_SIDE_PX
            current_max_side = max(img.width, img.height)
            if current_max_side > max_side:
                scale = max_side / float(current_max_side)
                new_size = (max(1, int(img.width * scale)), max(1, int(img.height * scale)))
                img = img.resize(new_size, resample)

            output = BytesIO()
            has_alpha = (
                "A" in img.getbands()
                or (img.mode == "P" and "transparency" in img.info)
            )
            if has_alpha:
                if img.mode != "RGBA":
                    img = img.convert("RGBA")
                img.save(output, format="PNG", optimize=True)
                return output.getvalue(), ".png", ""

            if img.mode not in {"RGB", "L"}:
                img = img.convert("RGB")
            img.save(
                output,
                format="JPEG",
                quality=88,
                optimize=True,
                progressive=True,
            )
            return output.getvalue(), ".jpg", ""
    except Exception:
        return image_bytes, ext, ""


def build_excel_image(xl_image_cls, image_path: str):
    if xl_image_cls is None:
        return None

    abs_path = resolve_uploaded_image_path(image_path)
    if not abs_path:
        return None

    try:
        return xl_image_cls(abs_path)
    except Exception:
        try:
            with open(abs_path, "rb") as fp:
                original_bytes = fp.read()
            ext = os.path.splitext(abs_path.lower())[1]
            image_for_excel, _, _ = polish_image_bytes(original_bytes, ext, logo=False)
            stream = BytesIO(image_for_excel)
            stream.seek(0)
            return xl_image_cls(stream)
        except Exception:
            return None


def save_requisicao_item_image(file_storage, clipboard_data: str):
    image_bytes, ext, erro = parse_image_payload(
        file_storage,
        clipboard_data,
        REQUISICAO_IMAGE_EXTENSIONS,
    )
    if erro:
        return "", erro
    if not image_bytes:
        return "", ""

    image_bytes, ext, _ = polish_image_bytes(image_bytes, ext, logo=False)
    if len(image_bytes) > MAX_IMAGE_SIZE_BYTES:
        return "", "Imagem muito grande. Limite de 8MB."

    upload_dir = os.path.join(app.root_path, "static", "uploads", "requisicoes")
    os.makedirs(upload_dir, exist_ok=True)
    filename = f"item_{uuid4().hex}{ext}"
    rel_path = f"uploads/requisicoes/{filename}"
    abs_path = os.path.join(app.root_path, "static", rel_path.replace("/", os.sep))
    with open(abs_path, "wb") as fp:
        fp.write(image_bytes)
    return rel_path, ""


def fetch_user_by_email(conn, email: str):
    normalized = normalize_email(email)
    if not normalized:
        return None
    return conn.execute(
        """
        SELECT
            id,
            nome,
            email,
            password_hash,
            provider,
            provider_sub,
            is_admin,
            is_active,
            created_em,
            updated_em,
            last_login_em,
            last_seen_em
        FROM users
        WHERE lower(email) = lower(?)
        LIMIT 1
        """,
        (normalized,),
    ).fetchone()


def fetch_user_by_provider_sub(conn, provider: str, provider_sub: str):
    provider_code = (provider or "").strip()
    sub = (provider_sub or "").strip()
    if not provider_code or not sub:
        return None
    return conn.execute(
        """
        SELECT
            id,
            nome,
            email,
            password_hash,
            provider,
            provider_sub,
            is_admin,
            is_active,
            created_em,
            updated_em,
            last_login_em,
            last_seen_em
        FROM users
        WHERE provider = ? AND provider_sub = ?
        LIMIT 1
        """,
        (provider_code, sub),
    ).fetchone()


def normalize_user_row(row):
    if not row:
        return None
    user = dict(row)
    user["is_admin"] = bool(user.get("is_admin"))
    user["is_active"] = bool(user.get("is_active"))
    return user


def load_current_user():
    user_id = session.get("user_id")
    user = fetch_user_by_id(user_id)
    if not user:
        logout_user()
        g.user = None
        return None
    if not user.get("is_active"):
        logout_user()
        g.user = None
        return None
    touch_user_activity(user)
    g.user = user
    return user


def build_user_activity_labels(user_row, now_dt=None):
    now_ref = now_dt or datetime.now()
    last_seen_dt = parse_iso_datetime(user_row.get("last_seen_em") or "")
    last_login_dt = parse_iso_datetime(user_row.get("last_login_em") or "")

    if last_seen_dt:
        seen_delta = now_ref - last_seen_dt
        if seen_delta.total_seconds() < USER_ONLINE_WINDOW_SECONDS:
            user_row["atividade_status"] = "online"
            user_row["atividade_tempo"] = f"Online ha {humanize_elapsed(seen_delta)}"
        else:
            user_row["atividade_status"] = "offline"
            user_row["atividade_tempo"] = f"Visto ha {humanize_elapsed(seen_delta)}"
    else:
        user_row["atividade_status"] = "offline"
        user_row["atividade_tempo"] = "Sem atividade"
    user_row["atividade_data"] = format_dt_pt(last_seen_dt)

    if last_login_dt:
        login_delta = now_ref - last_login_dt
        user_row["ultimo_login_tempo"] = f"Ha {humanize_elapsed(login_delta)}"
    else:
        user_row["ultimo_login_tempo"] = "Nunca"
    user_row["ultimo_login_data"] = format_dt_pt(last_login_dt)
    return user_row


@app.before_request
def enforce_auth_access():
    endpoint = request.endpoint or ""

    if endpoint.startswith("static"):
        return None

    user = load_current_user()

    if endpoint in PUBLIC_ENDPOINTS:
        if endpoint == "signup" and not AUTH_ALLOW_PUBLIC_SIGNUP and not user:
            return redirect(url_for("login", msg="Cadastro publico desativado."))
        if user and endpoint in {"login", "signup"}:
            return redirect(url_for("admin" if user.get("is_admin") else "index"))
        return None

    if not endpoint:
        return None

    if not user:
        login_url = url_for("login")
        next_url = current_request_relative_url()
        if is_safe_next_url(next_url):
            login_url = f"{login_url}?{urlencode({'next': next_url})}"
        return redirect(login_url)

    if endpoint.startswith("admin") and not user.get("is_admin"):
        return redirect(url_for("index", msg="Acesso restrito ao administrador."))

    return None


@app.context_processor
def inject_logged_user():
    user = getattr(g, "user", None)
    return {
        "current_user": user,
        "is_admin_user": bool(user and user.get("is_admin")),
        "auth_allow_public_signup": AUTH_ALLOW_PUBLIC_SIGNUP,
    }


@app.route("/login", methods=["GET", "POST"])
def login():
    next_url = pick_next_url(request.args.get("next", ""), request.form.get("next", ""))
    msg = (request.args.get("msg") or "").strip()
    msg_type = (request.args.get("msg_type") or "").strip()

    if request.method == "POST":
        email = normalize_email(request.form.get("email"))
        password = request.form.get("password") or ""

        if not email or not password:
            msg = "Informe email e senha."
            msg_type = "error"
        else:
            with closing(get_db()) as conn:
                row = fetch_user_by_email(conn, email)
                user = normalize_user_row(row)
                if not user or not user.get("password_hash"):
                    msg = "Credenciais invalidas."
                    msg_type = "error"
                elif not user.get("is_active"):
                    msg = "Usuario inativo. Fale com o administrador."
                    msg_type = "error"
                elif not check_password_hash(user["password_hash"], password):
                    msg = "Credenciais invalidas."
                    msg_type = "error"
                else:
                    ts = now_iso()
                    conn.execute(
                        """
                        UPDATE users
                        SET last_login_em = ?, last_seen_em = ?, updated_em = ?
                        WHERE id = ?
                        """,
                        (ts, ts, ts, user["id"]),
                    )
                    conn.commit()
                    login_user(user["id"])
                    return resolve_post_login_target(user)

    return render_template(
        "login.html",
        title=f"{APP_TITLE} - Acesso",
        msg=msg,
        msg_type=msg_type,
        next_url=next_url,
        allow_signup=AUTH_ALLOW_PUBLIC_SIGNUP,
        google_auth_enabled=is_google_auth_enabled(),
    )


@app.route("/esqueci-senha", methods=["GET", "POST"])
def forgot_password():
    next_url = pick_next_url(request.args.get("next", ""), request.form.get("next", ""))
    msg = (request.args.get("msg") or "").strip()
    msg_type = (request.args.get("msg_type") or "").strip()

    if request.method == "POST":
        email = normalize_email(request.form.get("email"))
        if "@" not in email:
            msg = "Informe um email valido."
            msg_type = "error"
        elif not is_password_reset_email_ready():
            msg = (
                "Recuperacao por email ainda nao configurada. "
                "Fale com o administrador para redefinir sua senha."
            )
            msg_type = "error"
        else:
            with closing(get_db()) as conn:
                user_row = fetch_user_by_email(conn, email)
                user = normalize_user_row(user_row)
                if user and user.get("is_active"):
                    raw_token = token_urlsafe(36)
                    token_hash = build_reset_token_hash(raw_token)
                    created_em = now_iso()
                    expires_em = (
                        datetime.now() + timedelta(minutes=PASSWORD_RESET_TOKEN_MINUTES)
                    ).isoformat()
                    conn.execute(
                        """
                        UPDATE password_reset_tokens
                        SET used_em = ?
                        WHERE user_id = ? AND used_em IS NULL
                        """,
                        (created_em, user["id"]),
                    )
                    conn.execute(
                        """
                        INSERT INTO password_reset_tokens
                            (user_id, token_hash, created_em, expires_em, used_em, requested_ip, requested_ua)
                        VALUES (?, ?, ?, ?, NULL, ?, ?)
                        """,
                        (
                            user["id"],
                            token_hash,
                            created_em,
                            expires_em,
                            (request.remote_addr or "")[:90],
                            (request.headers.get("User-Agent") or "")[:240],
                        ),
                    )
                    conn.commit()

                    reset_link = url_for("reset_password", token=raw_token, _external=True)
                    sent = send_password_reset_email(email, user.get("nome") or "", reset_link)
                    if not sent:
                        conn.execute(
                            "DELETE FROM password_reset_tokens WHERE token_hash = ?",
                            (token_hash,),
                        )
                        conn.commit()

            msg = (
                "Se o email estiver cadastrado, enviaremos um link de redefinicao em instantes."
            )
            msg_type = "success"

    return render_template(
        "forgot_password.html",
        title=f"{APP_TITLE} - Recuperar acesso",
        msg=msg,
        msg_type=msg_type,
        next_url=next_url,
        allow_signup=AUTH_ALLOW_PUBLIC_SIGNUP,
    )


@app.route("/reset-senha", methods=["GET", "POST"])
def reset_password():
    token = (request.args.get("token") or request.form.get("token") or "").strip()
    msg = (request.args.get("msg") or "").strip()
    msg_type = (request.args.get("msg_type") or "").strip()
    token_valid = False

    if request.method == "POST":
        senha = request.form.get("senha") or ""
        senha_confirm = request.form.get("senha_confirm") or ""
        if not token:
            msg = "Link invalido."
            msg_type = "error"
        elif len(senha) < 6:
            msg = "A senha deve ter pelo menos 6 caracteres."
            msg_type = "error"
        elif senha != senha_confirm:
            msg = "As senhas nao conferem."
            msg_type = "error"
        else:
            with closing(get_db()) as conn:
                token_row = fetch_valid_password_reset_token(conn, token)
                if not token_row:
                    msg = "Este link e invalido ou expirou. Solicite um novo."
                    msg_type = "error"
                else:
                    now_value = now_iso()
                    conn.execute(
                        """
                        UPDATE users
                        SET password_hash = ?, password_view_cipher = ?, provider = COALESCE(NULLIF(provider, ''), 'local'), updated_em = ?
                        WHERE id = ?
                        """,
                        (
                            generate_password_hash(senha),
                            build_password_view_cipher(senha),
                            now_value,
                            token_row["user_id"],
                        ),
                    )
                    conn.execute(
                        """
                        UPDATE password_reset_tokens
                        SET used_em = ?
                        WHERE user_id = ? AND used_em IS NULL
                        """,
                        (now_value, token_row["user_id"]),
                    )
                    conn.commit()
                    return redirect(
                        url_for("login", msg="Senha redefinida com sucesso. Entre com a nova senha.")
                    )

    if token:
        with closing(get_db()) as conn:
            token_valid = fetch_valid_password_reset_token(conn, token) is not None
        if not token_valid and not msg:
            msg = "Este link e invalido ou expirou. Solicite um novo."
            msg_type = "error"

    return render_template(
        "reset_password.html",
        title=f"{APP_TITLE} - Redefinir senha",
        msg=msg,
        msg_type=msg_type,
        token=token,
        token_valid=token_valid,
    )


@app.route("/signup", methods=["GET", "POST"])
def signup():
    if not AUTH_ALLOW_PUBLIC_SIGNUP:
        return redirect(url_for("login", msg="Cadastro publico desativado."))

    next_url = pick_next_url(request.args.get("next", ""), request.form.get("next", ""))
    msg = (request.args.get("msg") or "").strip()

    if request.method == "POST":
        nome = (request.form.get("nome") or "").strip()
        email = normalize_email(request.form.get("email"))
        senha = request.form.get("senha") or ""
        senha_confirm = request.form.get("senha_confirm") or ""

        if not nome:
            msg = "Informe o nome."
        elif "@" not in email:
            msg = "Informe um email valido."
        elif len(senha) < 6:
            msg = "A senha deve ter pelo menos 6 caracteres."
        elif senha != senha_confirm:
            msg = "As senhas nao conferem."
        else:
            with closing(get_db()) as conn:
                existente = fetch_user_by_email(conn, email)
                if existente:
                    msg = "Ja existe um usuario com este email."
                else:
                    ts = now_iso()
                    conn.execute(
                        """
                        INSERT INTO users
                            (nome, email, password_hash, password_view_cipher, provider, provider_sub, is_admin, is_active, created_em, updated_em, last_login_em, last_seen_em)
                        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                        """,
                        (
                            nome,
                            email,
                            generate_password_hash(senha),
                            build_password_view_cipher(senha),
                            "local",
                            "",
                            0,
                            1,
                            ts,
                            ts,
                            ts,
                            ts,
                        ),
                    )
                    user_id = conn.execute("SELECT last_insert_rowid()").fetchone()[0]
                    conn.commit()
                    login_user(user_id)
                    user = fetch_user_by_id(user_id)
                    return resolve_post_login_target(user, fallback_endpoint="index")

    return render_template(
        "signup.html",
        title=f"{APP_TITLE} - Criar conta",
        msg=msg,
        next_url=next_url,
        google_auth_enabled=is_google_auth_enabled(),
    )


@app.route("/logout", methods=["POST", "GET"])
def logout():
    logout_user()
    return redirect(url_for("login", msg="Sessao encerrada."))


@app.route("/login/google")
def login_google():
    if not is_google_auth_enabled():
        return redirect(url_for("login", msg="Login Google nao configurado."))

    next_url = pick_next_url(request.args.get("next", ""))
    state = token_urlsafe(24)
    session["oauth_google_state"] = state
    session["oauth_google_next"] = next_url

    try:
        config = get_google_openid_config()
        auth_endpoint = config.get("authorization_endpoint")
        if not auth_endpoint:
            raise RuntimeError("authorization endpoint ausente")
        params = {
            "client_id": GOOGLE_CLIENT_ID,
            "redirect_uri": url_for("google_callback", _external=True),
            "response_type": "code",
            "scope": "openid email profile",
            "state": state,
            "prompt": "select_account",
        }
        return redirect(f"{auth_endpoint}?{urlencode(params)}")
    except Exception:
        return redirect(url_for("login", msg="Falha ao iniciar login Google."))


@app.route("/auth/google/callback")
def google_callback():
    if not is_google_auth_enabled():
        return redirect(url_for("login", msg="Login Google nao configurado."))

    incoming_state = (request.args.get("state") or "").strip()
    expected_state = (session.pop("oauth_google_state", "") or "").strip()
    if not incoming_state or incoming_state != expected_state:
        return redirect(url_for("login", msg="Sessao Google invalida. Tente novamente."))

    auth_code = (request.args.get("code") or "").strip()
    if not auth_code:
        return redirect(url_for("login", msg="Nao foi possivel concluir o login Google."))

    try:
        config = get_google_openid_config()
        token_endpoint = config.get("token_endpoint")
        userinfo_endpoint = config.get("userinfo_endpoint")
        if not token_endpoint or not userinfo_endpoint:
            raise RuntimeError("configuracao OpenID incompleta")

        token_payload = urlencode(
            {
                "code": auth_code,
                "client_id": GOOGLE_CLIENT_ID,
                "client_secret": GOOGLE_CLIENT_SECRET,
                "redirect_uri": url_for("google_callback", _external=True),
                "grant_type": "authorization_code",
            }
        ).encode("utf-8")

        token_request = Request(
            token_endpoint,
            data=token_payload,
            headers={
                "Content-Type": "application/x-www-form-urlencoded",
                "Accept": "application/json",
            },
        )
        with urlopen(token_request, timeout=8) as token_response:
            token_data = json.loads(token_response.read().decode("utf-8"))

        access_token = (token_data.get("access_token") or "").strip()
        if not access_token:
            raise RuntimeError("access token ausente")

        userinfo_request = Request(
            userinfo_endpoint,
            headers={
                "Authorization": f"Bearer {access_token}",
                "Accept": "application/json",
            },
        )
        with urlopen(userinfo_request, timeout=8) as userinfo_response:
            profile = json.loads(userinfo_response.read().decode("utf-8"))

        provider_sub = (profile.get("sub") or "").strip()
        email = normalize_email(profile.get("email") or "")
        nome = (profile.get("name") or "").strip()
        if not provider_sub or not email:
            raise RuntimeError("perfil Google sem email/sub")

        with closing(get_db()) as conn:
            user_row = fetch_user_by_provider_sub(conn, "google", provider_sub)
            if not user_row:
                user_row = fetch_user_by_email(conn, email)

            ts = now_iso()
            if user_row:
                user = normalize_user_row(user_row)
                if not user.get("is_active"):
                    return redirect(url_for("login", msg="Usuario inativo. Fale com o administrador."))

                updates = ["last_login_em = ?", "last_seen_em = ?", "updated_em = ?"]
                params = [ts, ts, ts]
                if not (user.get("provider_sub") or "").strip():
                    updates.append("provider_sub = ?")
                    params.append(provider_sub)
                if not (user.get("provider") or "").strip():
                    updates.append("provider = ?")
                    params.append("google")
                if nome and not (user.get("nome") or "").strip():
                    updates.append("nome = ?")
                    params.append(nome)
                params.append(user["id"])
                conn.execute(f"UPDATE users SET {', '.join(updates)} WHERE id = ?", params)
                conn.commit()
                login_user(user["id"])
                return resolve_post_login_target(user)

            nome_final = nome or email.split("@")[0]
            conn.execute(
                """
                INSERT INTO users
                    (nome, email, password_hash, provider, provider_sub, is_admin, is_active, created_em, updated_em, last_login_em, last_seen_em)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                """,
                (
                    nome_final,
                    email,
                    None,
                    "google",
                    provider_sub,
                    0,
                    1,
                    ts,
                    ts,
                    ts,
                    ts,
                ),
            )
            user_id = conn.execute("SELECT last_insert_rowid()").fetchone()[0]
            conn.commit()
            login_user(user_id)
            user = fetch_user_by_id(user_id)
            return resolve_post_login_target(user, fallback_endpoint="index")
    except (HTTPError, URLError, ValueError, RuntimeError):
        return redirect(url_for("login", msg="Falha no login Google."))


@app.route("/")
def index():
    msg = request.args.get("msg", "")
    hoje = datetime.now().date()
    default_necessario = build_necessario_min_date(0, hoje)
    with closing(get_db()) as conn:
        cur = conn.cursor()
        obras = rows_to_dicts(
            cur.execute("SELECT id, nome FROM obras ORDER BY nome").fetchall()
        )
        obras = filter_obras_for_user(cur, g.get("user"), obras)
        categorias = rows_to_dicts(
            cur.execute("SELECT id, nome FROM categorias ORDER BY nome").fetchall()
        )
        insumos = rows_to_dicts(
            cur.execute(
                """
                SELECT
                    id,
                    TRIM(nome) AS nome,
                    categoria_id,
                    foto_path
                FROM insumos
                WHERE TRIM(COALESCE(nome, '')) <> ''
                  AND COALESCE(is_custom, 0) = 0
                ORDER BY TRIM(nome)
                """
            ).fetchall()
        )
        specs = rows_to_dicts(
            cur.execute(
                "SELECT id, insumo_id, nome FROM especificacoes ORDER BY nome"
            ).fetchall()
        )
        obra_ids = [int(obra["id"]) for obra in obras]
        if obra_ids:
            placeholders = ",".join(["?"] * len(obra_ids))
            apropriacoes = rows_to_dicts(
                cur.execute(
                    f"""
                    SELECT
                        id,
                        obra_id,
                        numero,
                        TRIM(COALESCE(nome, '')) AS nome
                    FROM apropriacoes
                    WHERE obra_id IN ({placeholders})
                    ORDER BY COALESCE(NULLIF(TRIM(nome), ''), numero), numero
                    """,
                    obra_ids,
                ).fetchall()
            )
        else:
            apropriacoes = []
        unidades = rows_to_dicts(
            cur.execute("SELECT id, nome FROM unidades ORDER BY nome").fetchall()
        )
        insumo_unidades = rows_to_dicts(
            cur.execute(
                """
                SELECT iu.insumo_id, u.nome
                FROM insumo_unidades iu
                JOIN unidades u ON u.id = iu.unidade_id
                JOIN insumos i ON i.id = iu.insumo_id
                WHERE COALESCE(i.is_custom, 0) = 0
                ORDER BY u.nome
                """
            ).fetchall()
        )
    if not obras and not msg:
        msg = "Nenhuma obra liberada para seu usuario. Fale com o administrador."
    return render_template(
        "index.html",
        title=APP_TITLE,
        obras=obras,
        categorias=categorias,
        insumos=insumos,
        specs=specs,
        apropriacoes=apropriacoes,
        unidades=unidades,
        insumo_unidades=insumo_unidades,
        today_iso=hoje.isoformat(),
        default_necessario_iso=default_necessario.isoformat(),
        lead_rules_payload=[
            {"max_items": max_items, "dias": days}
            for max_items, days in REQUEST_LEAD_RULES
        ],
        lead_fallback_days=REQUEST_LEAD_FALLBACK_DAYS,
        msg=msg,
        default_solicitante=((g.user or {}).get("nome") or ""),
    )


@app.route("/requisicao", methods=["POST"])
def criar_requisicao():
    obra_id = request.form.get("obra")
    solicitante = (request.form.get("solicitante") or "").strip()
    if not solicitante and g.get("user"):
        solicitante = (g.user.get("nome") or "").strip()
    necessario_em = (request.form.get("necessario_em") or "").strip()
    if necessario_em and not re.fullmatch(r"\d{4}-\d{2}-\d{2}", necessario_em):
        necessario_em = ""
    if not obra_id:
        return redirect(url_for("index"))
    try:
        obra_id_int = int(obra_id)
    except (TypeError, ValueError):
        return redirect(url_for("index", msg="Obra invalida."))

    with closing(get_db()) as conn:
        cur = conn.cursor()
        if not user_has_access_to_obra(cur, g.get("user"), obra_id_int):
            return redirect(
                url_for("index", msg="Voce nao possui acesso a esta obra.")
            )

    apropriacao_nome_lookup = build_apropriacao_nome_lookup(obra_id_int)

    itens = []
    index_pattern = re.compile(r"^itens\[(\d+)\]\[insumo\]$")
    indices = sorted(
        {
            int(match.group(1))
            for key in request.form.keys()
            for match in [index_pattern.match(key)]
            if match
        }
    )
    for idx in indices:
        insumo_id_raw = request.form.get(f"itens[{idx}][insumo]")
        insumo_custom = (request.form.get(f"itens[{idx}][insumo_custom]") or "").strip()
        unidade = request.form.get(f"itens[{idx}][unidade]") or ""
        quantidade = request.form.get(f"itens[{idx}][quantidade]") or "0"
        especificacao = request.form.get(f"itens[{idx}][especificacao]") or ""
        apropriacao_raw = request.form.get(f"itens[{idx}][apropriacao]") or ""
        item_necessario_em = (request.form.get(f"itens[{idx}][necessario_em]") or "").strip()
        link_produto = (request.form.get(f"itens[{idx}][link_produto]") or "").strip()
        foto_path, foto_erro = save_requisicao_item_image(
            request.files.get(f"itens[{idx}][foto]"),
            request.form.get(f"itens[{idx}][foto_clipboard]") or "",
        )
        if foto_erro:
            return redirect(url_for("index", msg=foto_erro))
        if item_necessario_em and not re.fullmatch(r"\d{4}-\d{2}-\d{2}", item_necessario_em):
            item_necessario_em = ""
        apropriacao = resolve_apropriacao_codigo(
            apropriacao_raw, apropriacao_nome_lookup
        )
        insumo_id = None
        if insumo_id_raw:
            try:
                insumo_id = int(insumo_id_raw)
            except (TypeError, ValueError):
                insumo_id = None

        tem_insumo = bool(insumo_id) or bool(insumo_custom)
        if tem_insumo and unidade:
            try:
                quantidade_val = float(quantidade.replace(",", "."))
            except ValueError:
                quantidade_val = 0.0
            itens.append(
                {
                    "insumo_id": insumo_id,
                    "insumo_custom": insumo_custom,
                    "is_custom": bool(insumo_custom and not insumo_id),
                    "unidade": unidade,
                    "quantidade": quantidade_val,
                    "especificacao": especificacao,
                    "apropriacao": apropriacao,
                    "necessario_em": item_necessario_em or necessario_em,
                    "link_produto": link_produto,
                    "foto_path": foto_path,
                }
            )

    if not itens:
        return redirect(url_for("index"))

    min_date = build_necessario_min_date(len(itens))
    min_label = min_date.strftime("%d/%m/%Y")
    min_iso = min_date.isoformat()

    base_necessario = parse_iso_date(necessario_em)
    if base_necessario is None:
        necessario_em = min_iso
    elif base_necessario < min_date:
        return redirect(
            url_for(
                "index",
                msg=(
                    f"Para {len(itens)} insumo(s), a data minima permitida de necessidade e {min_label}."
                ),
            )
        )

    for item in itens:
        item_date = parse_iso_date(item.get("necessario_em", ""))
        if item_date is None:
            item["necessario_em"] = necessario_em
            item_date = parse_iso_date(item["necessario_em"])
        if item_date is None or item_date < min_date:
            return redirect(
                url_for(
                    "index",
                    msg=(
                        f"Para {len(itens)} insumo(s), a data minima permitida de necessidade e {min_label}."
                    ),
                )
            )

    def ensure_custom_insumo(cur, nome: str) -> int:
        nome_limpo = (nome or "").strip()
        if not nome_limpo:
            return 0
        existente = cur.execute(
            """
            SELECT id
            FROM insumos
            WHERE COALESCE(is_custom, 0) = 1
              AND lower(trim(nome)) = lower(trim(?))
            LIMIT 1
            """,
            (nome_limpo,),
        ).fetchone()
        if existente:
            return int(existente["id"])
        cur.execute("INSERT INTO insumos (nome, is_custom) VALUES (?, 1)", (nome_limpo,))
        return int(cur.lastrowid)

    with closing(get_db()) as conn:
        cur = conn.cursor()
        cur.execute(
            """
            INSERT INTO requisicoes (obra_id, criado_em, solicitante, necessario_em)
            VALUES (?, ?, ?, ?)
            """,
            (obra_id_int, datetime.now().isoformat(), solicitante, necessario_em),
        )
        requisicao_id = cur.lastrowid
        for item in itens:
            insumo_id = item.get("insumo_id")
            if item.get("is_custom"):
                insumo_id = ensure_custom_insumo(cur, item.get("insumo_custom"))
                item["insumo_id"] = insumo_id
            if not insumo_id:
                return redirect(
                    url_for(
                        "index",
                        msg="Nao foi possivel registrar um dos insumos informados.",
                    )
                )
            cur.execute(
                """
                INSERT INTO requisicao_itens
                    (
                        requisicao_id,
                        insumo_id,
                        unidade,
                        quantidade,
                        especificacao,
                        apropriacao,
                        necessario_em,
                        link_produto,
                        foto_path
                    )
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
                """,
                (
                    requisicao_id,
                    insumo_id,
                    item["unidade"],
                    item["quantidade"],
                    item["especificacao"],
                    item["apropriacao"],
                    item["necessario_em"],
                    item["link_produto"],
                    item["foto_path"],
                ),
            )
        conn.commit()

    return redirect(url_for("exportar_requisicao", requisicao_id=requisicao_id))


@app.route("/requisicao/<int:requisicao_id>/export")
def exportar_requisicao(requisicao_id: int):
    try:
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import A4, landscape
        from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
        from reportlab.lib.units import mm
        from reportlab.platypus import (
            Image as PdfImage,
            Paragraph,
            SimpleDocTemplate,
            Spacer,
            Table,
            TableStyle,
        )
        from xml.sax.saxutils import escape as xml_escape
    except Exception:
        return "Instale reportlab: pip install reportlab", 500

    unidades_por_insumo = {}
    with closing(get_db()) as conn:
        cab = conn.execute(
            """
            SELECT r.id, r.criado_em, r.solicitante, r.necessario_em, o.nome as obra
            FROM requisicoes r
            JOIN obras o ON o.id = r.obra_id
            WHERE r.id = ?
            """,
            (requisicao_id,),
        ).fetchone()

        if not cab:
            return "Requisicao nao encontrada.", 404

        itens = conn.execute(
            """
            SELECT
                ri.insumo_id,
                i.nome as insumo,
                COALESCE(i.is_custom, 0) as is_custom,
                ri.unidade,
                ri.quantidade,
                ri.especificacao,
                ri.apropriacao,
                ri.necessario_em,
                ri.link_produto,
                ri.foto_path
            FROM requisicao_itens ri
            JOIN insumos i ON i.id = ri.insumo_id
            WHERE ri.requisicao_id = ?
            ORDER BY ri.id
            """,
            (requisicao_id,),
        ).fetchall()

        insumo_ids = sorted(
            {
                int(row["insumo_id"])
                for row in itens
                if row is not None and row["insumo_id"] is not None
            }
        )
        if insumo_ids:
            placeholders = ",".join(["?"] * len(insumo_ids))
            unidades_rows = conn.execute(
                f"""
                SELECT iu.insumo_id, u.nome
                FROM insumo_unidades iu
                JOIN unidades u ON u.id = iu.unidade_id
                WHERE iu.insumo_id IN ({placeholders})
                """,
                insumo_ids,
            ).fetchall()
            for row in unidades_rows:
                insumo_id = int(row["insumo_id"] or 0)
                if not insumo_id:
                    continue
                unidades_por_insumo.setdefault(insumo_id, set()).add(
                    normalize_text(row["nome"])
                )

    def safe_text(value, default="-") -> str:
        raw = str(value or "").strip()
        if not raw:
            return default
        return xml_escape(raw).replace("\n", "<br/>")

    def format_quantidade(value) -> str:
        try:
            num = float(value)
        except (TypeError, ValueError):
            return str(value or "")
        if num.is_integer():
            return str(int(num))
        return (f"{num:.4f}").rstrip("0").rstrip(".").replace(".", ",")

    def build_pdf_media_cell(link_produto: str, foto_path: str):
        blocos = []
        link_texto = (link_produto or "").strip()
        if link_texto:
            blocos.append(Paragraph(safe_text(link_texto), media_text_style))

        foto_raw = (foto_path or "").strip()
        foto_carregada = False
        if foto_raw:
            abs_path = resolve_uploaded_image_path(foto_raw)
            if abs_path:
                foto_pdf = None
                try:
                    foto_pdf = PdfImage(abs_path)
                except Exception:
                    try:
                        with open(abs_path, "rb") as fp:
                            original_bytes = fp.read()
                        ext = os.path.splitext(abs_path.lower())[1]
                        image_for_pdf, _, _ = polish_image_bytes(
                            original_bytes,
                            ext,
                            logo=False,
                        )
                        stream = BytesIO(image_for_pdf)
                        stream.seek(0)
                        foto_pdf = PdfImage(stream)
                    except Exception:
                        foto_pdf = None

                if foto_pdf:
                    max_width = 66 * mm
                    max_height = 24 * mm
                    width = float(getattr(foto_pdf, "imageWidth", 0) or 0)
                    height = float(getattr(foto_pdf, "imageHeight", 0) or 0)
                    if width > 0 and height > 0:
                        escala = min(max_width / width, max_height / height, 1.0)
                        foto_pdf.drawWidth = width * escala
                        foto_pdf.drawHeight = height * escala
                    else:
                        foto_pdf.drawWidth = max_width
                        foto_pdf.drawHeight = max_height

                    if blocos:
                        blocos.append(Spacer(1, 1 * mm))
                    blocos.append(foto_pdf)
                    foto_carregada = True

        if foto_raw and not foto_carregada:
            if blocos:
                blocos.append(Spacer(1, 1 * mm))
            blocos.append(Paragraph("Foto anexada (erro ao carregar).", media_text_style))

        if not blocos:
            return Paragraph("-", normal_style)
        return blocos

    output = BytesIO()
    filename = f"requisicao_{requisicao_id}.pdf"

    doc = SimpleDocTemplate(
        output,
        pagesize=landscape(A4),
        leftMargin=9 * mm,
        rightMargin=9 * mm,
        topMargin=10 * mm,
        bottomMargin=10 * mm,
        title=f"Requisicao {requisicao_id}",
    )

    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        "PdfTitle",
        parent=styles["Heading1"],
        fontName="Helvetica-Bold",
        fontSize=14,
        leading=17,
        alignment=1,
        spaceAfter=6,
    )
    normal_style = ParagraphStyle(
        "PdfNormal",
        parent=styles["BodyText"],
        fontName="Helvetica",
        fontSize=8,
        leading=10,
    )
    media_text_style = ParagraphStyle(
        "PdfMedia",
        parent=normal_style,
        fontSize=7.5,
        leading=9,
        wordWrap="CJK",
    )
    label_style = ParagraphStyle(
        "PdfLabel",
        parent=normal_style,
        fontName="Helvetica-Bold",
        fontSize=8,
    )
    unidade_alert_style = ParagraphStyle(
        "PdfUnidadeAlert",
        parent=normal_style,
        textColor=colors.HexColor("#C62828"),
    )
    custom_insumo_style = ParagraphStyle(
        "PdfCustomInsumo",
        parent=normal_style,
        textColor=colors.HexColor("#C62828"),
    )

    def is_unidade_predefinida(insumo_id: int, unidade_nome: str) -> bool:
        if not insumo_id:
            return False
        nomes = unidades_por_insumo.get(int(insumo_id))
        if not nomes:
            return False
        return normalize_text(unidade_nome) in nomes

    story = []
    story.append(Paragraph("FICHA PARA PEDIDO DE COMPRA", title_style))
    story.append(Paragraph("CONSTRUTORA SUL CAPIXABA LTDA", label_style))
    story.append(Spacer(1, 4 * mm))

    info_data = [
        [
            Paragraph("<b>OBRA</b>", label_style),
            Paragraph(safe_text(cab["obra"]), normal_style),
            Paragraph("<b>SOLICITANTE</b>", label_style),
            Paragraph(safe_text(cab["solicitante"]), normal_style),
        ],
        [
            Paragraph("<b>DATA SOLICITACAO</b>", label_style),
            Paragraph(safe_text(format_date_for_sheet(cab["criado_em"])), normal_style),
            Paragraph("<b>NECESSARIO PARA</b>", label_style),
            Paragraph(
                safe_text(format_date_for_sheet(cab["necessario_em"], cab["criado_em"])),
                normal_style,
            ),
        ],
    ]

    info_table = Table(
        info_data,
        colWidths=[26 * mm, 102 * mm, 36 * mm, 115 * mm],
    )
    info_table.setStyle(
        TableStyle(
            [
                ("GRID", (0, 0), (-1, -1), 0.5, colors.black),
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                ("BACKGROUND", (0, 0), (0, -1), colors.HexColor("#ECECEC")),
                ("BACKGROUND", (2, 0), (2, -1), colors.HexColor("#ECECEC")),
                ("LEFTPADDING", (0, 0), (-1, -1), 5),
                ("RIGHTPADDING", (0, 0), (-1, -1), 5),
                ("TOPPADDING", (0, 0), (-1, -1), 4),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
            ]
        )
    )
    story.append(info_table)
    story.append(Spacer(1, 5 * mm))

    header_labels = [
        "ITEM",
        "INSUMO",
        "UNIDADE",
        "QTD",
        "ESPECIFICACAO",
        "APROPRIACAO",
        "NECESSARIO",
        "LINK/FOTO",
    ]

    table_data = [
        [Paragraph(f"<b>{safe_text(label, default='')}</b>", label_style) for label in header_labels]
    ]

    for idx, item in enumerate(itens, start=1):
        necessario_item = format_date_for_sheet(
            item["necessario_em"],
            cab["necessario_em"] or cab["criado_em"],
        )
        media_cell = build_pdf_media_cell(item["link_produto"], item["foto_path"])
        unidade_style = (
            normal_style
            if is_unidade_predefinida(item["insumo_id"], item["unidade"])
            else unidade_alert_style
        )
        insumo_style = custom_insumo_style if bool(item["is_custom"]) else normal_style

        table_data.append(
            [
                Paragraph(str(idx), normal_style),
                Paragraph(safe_text(item["insumo"]), insumo_style),
                Paragraph(safe_text(item["unidade"]), unidade_style),
                Paragraph(safe_text(format_quantidade(item["quantidade"])), normal_style),
                Paragraph(safe_text(item["especificacao"]), normal_style),
                Paragraph(safe_text(item["apropriacao"]), normal_style),
                Paragraph(safe_text(necessario_item), normal_style),
                media_cell,
            ]
        )

    itens_table = Table(
        table_data,
        colWidths=[12 * mm, 55 * mm, 16 * mm, 18 * mm, 55 * mm, 28 * mm, 23 * mm, 72 * mm],
        repeatRows=1,
    )
    itens_table.setStyle(
        TableStyle(
            [
                ("GRID", (0, 0), (-1, -1), 0.45, colors.black),
                ("VALIGN", (0, 0), (-1, -1), "TOP"),
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#D5DEEC")),
                ("ALIGN", (0, 0), (0, -1), "CENTER"),
                ("ALIGN", (2, 1), (3, -1), "CENTER"),
                ("ALIGN", (5, 1), (6, -1), "CENTER"),
                ("LEFTPADDING", (0, 0), (-1, -1), 4),
                ("RIGHTPADDING", (0, 0), (-1, -1), 4),
                ("TOPPADDING", (0, 0), (-1, -1), 3),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
            ]
        )
    )
    story.append(itens_table)
    story.append(Spacer(1, 4 * mm))

    story.append(
        Paragraph(
            (
                "<b>Observacoes importantes:</b> Solicite insumos com dados completos e claros. "
                "A apropriacao deve estar correta para a obra."
            ),
            normal_style,
        )
    )

    doc.build(story)
    output.seek(0)
    return send_file(
        output,
        as_attachment=True,
        download_name=filename,
        mimetype="application/pdf",
    )


@app.route("/admin")
def admin():
    import_result = request.args.get("import_result", "")
    admin_msg = request.args.get("admin_msg", "")
    return render_template(
        "admin.html",
        title=APP_TITLE,
        current_admin="overview",
        import_result=import_result,
        admin_msg=admin_msg,
    )


@app.route("/admin/usuarios", methods=["GET", "POST"])
def admin_usuarios():
    mensagem = (request.args.get("msg") or "").strip()

    if request.method == "POST":
        nome = (request.form.get("nome") or "").strip()
        email = normalize_email(request.form.get("email"))
        senha = request.form.get("senha") or ""
        is_admin = request.form.get("is_admin") == "1"
        is_active = request.form.get("is_active") != "0"

        if not nome:
            mensagem = "Informe o nome."
        elif "@" not in email:
            mensagem = "Informe um email valido."
        elif len(senha) < 6:
            mensagem = "A senha deve ter pelo menos 6 caracteres."
        else:
            with closing(get_db()) as conn:
                existente = fetch_user_by_email(conn, email)
                if existente:
                    mensagem = "Ja existe um usuario com este email."
                else:
                    ts = now_iso()
                    conn.execute(
                        """
                        INSERT INTO users
                            (nome, email, password_hash, password_view_cipher, provider, provider_sub, is_admin, is_active, created_em, updated_em, last_login_em, last_seen_em)
                        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                        """,
                        (
                            nome,
                            email,
                            generate_password_hash(senha),
                            build_password_view_cipher(senha),
                            "local",
                            "",
                            1 if is_admin else 0,
                            1 if is_active else 0,
                            ts,
                            ts,
                            None,
                            None,
                        ),
                    )
                    conn.commit()
                    return redirect(url_for("admin_usuarios", msg="Usuario criado com sucesso."))

    with closing(get_db()) as conn:
        cur = conn.cursor()
        usuarios = rows_to_dicts(
            cur.execute(
                """
                SELECT
                    id,
                    nome,
                    email,
                    provider,
                    is_admin,
                    is_active,
                    created_em,
                    last_login_em,
                    last_seen_em
                FROM users
                ORDER BY is_admin DESC, is_active DESC, nome, email
                """
            ).fetchall()
        )
        obras = rows_to_dicts(cur.execute("SELECT id, nome FROM obras ORDER BY nome").fetchall())
        user_obra_rows = rows_to_dicts(
            cur.execute("SELECT user_id, obra_id FROM user_obras ORDER BY user_id, obra_id").fetchall()
        )

        user_obra_map = {}
        for row in user_obra_rows:
            user_id = int(row.get("user_id") or 0)
            obra_id = int(row.get("obra_id") or 0)
            if not user_id or not obra_id:
                continue
            user_obra_map.setdefault(user_id, []).append(obra_id)
        obra_name_map = {int(obra["id"]): obra["nome"] for obra in obras}

        now_ref = datetime.now()
        usuarios = [build_user_activity_labels(usuario, now_ref) for usuario in usuarios]
        for usuario in usuarios:
            usuario["is_master_admin"] = is_master_admin_email(usuario.get("email", ""))
            allowed_ids = user_obra_map.get(int(usuario.get("id") or 0), [])
            usuario["allowed_obra_ids"] = allowed_ids
            if usuario["is_master_admin"]:
                usuario["allowed_obras_label"] = "Todas as obras (master)"
            elif not allowed_ids:
                usuario["allowed_obras_label"] = "Nenhuma obra permitida"
            else:
                nomes = [obra_name_map.get(obra_id, "") for obra_id in allowed_ids]
                nomes = [nome for nome in nomes if nome]
                usuario["allowed_obras_label"] = ", ".join(nomes) if nomes else "Nenhuma obra valida"

    return render_template(
        "admin_usuarios.html",
        title=APP_TITLE,
        current_admin="usuarios",
        usuarios=usuarios,
        obras=obras,
        mensagem=mensagem,
    )


@app.route("/admin/usuarios/<int:user_id>/obras", methods=["POST"])
def admin_usuarios_obras(user_id: int):
    obra_ids = parse_int_list(request.form.getlist("obra_ids"))

    with closing(get_db()) as conn:
        cur = conn.cursor()
        user_row = cur.execute(
            "SELECT id, email FROM users WHERE id = ?",
            (user_id,),
        ).fetchone()
        if not user_row:
            return redirect(url_for("admin_usuarios", msg="Usuario nao encontrado."))

        if is_master_admin_row(user_row):
            return redirect(
                url_for(
                    "admin_usuarios",
                    msg="Administrador master possui acesso total e nao precisa de vinculacao de obras.",
                )
            )

        if obra_ids:
            placeholders = ",".join(["?"] * len(obra_ids))
            valid_rows = cur.execute(
                f"SELECT id FROM obras WHERE id IN ({placeholders})",
                obra_ids,
            ).fetchall()
            valid_ids = sorted({int(row["id"]) for row in valid_rows})
        else:
            valid_ids = []

        cur.execute("DELETE FROM user_obras WHERE user_id = ?", (user_id,))
        if valid_ids:
            ts = now_iso()
            cur.executemany(
                "INSERT INTO user_obras (user_id, obra_id, created_em) VALUES (?, ?, ?)",
                [(user_id, obra_id, ts) for obra_id in valid_ids],
            )
        conn.commit()

    if valid_ids:
        return redirect(
            url_for(
                "admin_usuarios",
                msg=f"Permissoes atualizadas: {len(valid_ids)} obra(s) vinculada(s).",
            )
        )
    return redirect(
        url_for(
            "admin_usuarios",
            msg="Permissoes atualizadas: usuario sem obras liberadas.",
        )
    )


@app.route("/admin/usuarios/<int:user_id>/toggle-admin", methods=["POST"])
def admin_usuarios_toggle_admin(user_id: int):
    current_user_id = int((g.user or {}).get("id") or 0)
    with closing(get_db()) as conn:
        cur = conn.cursor()
        user_row = cur.execute(
            "SELECT id, nome, email, is_admin, is_active FROM users WHERE id = ?",
            (user_id,),
        ).fetchone()
        if not user_row:
            return redirect(url_for("admin_usuarios", msg="Usuario nao encontrado."))

        user = dict(user_row)
        if is_master_admin_row(user):
            return redirect(
                url_for(
                    "admin_usuarios",
                    msg="O administrador master e protegido e nao pode perder o perfil admin.",
                )
            )
        is_admin = bool(user["is_admin"])
        next_admin = 0 if is_admin else 1

        if is_admin and next_admin == 0 and count_active_admins(cur) <= 1:
            return redirect(url_for("admin_usuarios", msg="Nao e possivel remover o ultimo admin ativo."))

        cur.execute(
            "UPDATE users SET is_admin = ?, updated_em = ? WHERE id = ?",
            (next_admin, now_iso(), user_id),
        )
        conn.commit()

    if current_user_id == user_id and next_admin == 0:
        logout_user()
        return redirect(url_for("login", msg="Seu perfil admin foi removido."))

    return redirect(url_for("admin_usuarios", msg="Perfil atualizado."))


@app.route("/admin/usuarios/<int:user_id>/toggle-active", methods=["POST"])
def admin_usuarios_toggle_active(user_id: int):
    current_user_id = int((g.user or {}).get("id") or 0)
    with closing(get_db()) as conn:
        cur = conn.cursor()
        user_row = cur.execute(
            "SELECT id, email, is_admin, is_active FROM users WHERE id = ?",
            (user_id,),
        ).fetchone()
        if not user_row:
            return redirect(url_for("admin_usuarios", msg="Usuario nao encontrado."))

        user = dict(user_row)
        if is_master_admin_row(user):
            return redirect(
                url_for(
                    "admin_usuarios",
                    msg="O administrador master e protegido e nao pode ser desativado.",
                )
            )
        is_active = bool(user["is_active"])
        next_active = 0 if is_active else 1

        if current_user_id == user_id and next_active == 0:
            return redirect(url_for("admin_usuarios", msg="Voce nao pode desativar seu proprio acesso."))

        if bool(user["is_admin"]) and next_active == 0 and count_active_admins(cur) <= 1:
            return redirect(url_for("admin_usuarios", msg="Nao e possivel desativar o ultimo admin ativo."))

        cur.execute(
            "UPDATE users SET is_active = ?, updated_em = ? WHERE id = ?",
            (next_active, now_iso(), user_id),
        )
        conn.commit()

    return redirect(url_for("admin_usuarios", msg="Status atualizado."))


@app.route("/admin/usuarios/<int:user_id>/senha", methods=["POST"])
def admin_usuarios_reset_senha(user_id: int):
    nova_senha = request.form.get("nova_senha") or ""
    if len(nova_senha) < 6:
        return redirect(url_for("admin_usuarios", msg="Senha invalida. Use pelo menos 6 caracteres."))

    with closing(get_db()) as conn:
        cur = conn.cursor()
        user_row = cur.execute(
            "SELECT id, email FROM users WHERE id = ?",
            (user_id,),
        ).fetchone()
        if not user_row:
            return redirect(url_for("admin_usuarios", msg="Usuario nao encontrado."))
        if is_master_admin_row(user_row):
            return redirect(
                url_for(
                    "admin_usuarios",
                    msg="A senha do administrador master e protegida por configuracao do sistema.",
                )
            )

        cur.execute(
            """
            UPDATE users
            SET password_hash = ?, password_view_cipher = ?, provider = COALESCE(NULLIF(provider, ''), 'local'), updated_em = ?
            WHERE id = ?
            """,
            (
                generate_password_hash(nova_senha),
                build_password_view_cipher(nova_senha),
                now_iso(),
                user_id,
            ),
        )
        conn.commit()

    return redirect(url_for("admin_usuarios", msg="Senha atualizada com sucesso."))


@app.route("/admin/usuarios/<int:user_id>/senha/view", methods=["GET"])
def admin_usuarios_view_senha(user_id: int):
    with closing(get_db()) as conn:
        row = conn.execute(
            "SELECT id, email, password_view_cipher FROM users WHERE id = ?",
            (user_id,),
        ).fetchone()
    if not row:
        return jsonify({"ok": False, "msg": "Usuario nao encontrado."}), 404

    senha = reveal_password_from_cipher((row["password_view_cipher"] or "").strip())
    if not senha:
        return jsonify(
            {
                "ok": False,
                "msg": "Senha indisponivel para exibicao. Redefina a senha para registrar uma nova.",
            }
        )
    return jsonify({"ok": True, "senha": senha})


@app.route("/admin/usuarios/<int:user_id>/delete", methods=["POST"])
def admin_usuarios_delete(user_id: int):
    current_user_id = int((g.user or {}).get("id") or 0)
    if current_user_id == user_id:
        return redirect(url_for("admin_usuarios", msg="Nao e permitido excluir seu proprio usuario."))

    with closing(get_db()) as conn:
        cur = conn.cursor()
        user_row = cur.execute(
            "SELECT id, email, is_admin, is_active FROM users WHERE id = ?",
            (user_id,),
        ).fetchone()
        if not user_row:
            return redirect(url_for("admin_usuarios", msg="Usuario nao encontrado."))

        user = dict(user_row)
        if is_master_admin_row(user):
            return redirect(
                url_for(
                    "admin_usuarios",
                    msg="O administrador master e protegido e nao pode ser excluido.",
                )
            )
        if bool(user["is_admin"]) and bool(user["is_active"]) and count_active_admins(cur) <= 1:
            return redirect(url_for("admin_usuarios", msg="Nao e possivel excluir o ultimo admin ativo."))

        cur.execute("DELETE FROM user_obras WHERE user_id = ?", (user_id,))
        cur.execute("DELETE FROM users WHERE id = ?", (user_id,))
        conn.commit()

    return redirect(url_for("admin_usuarios", msg="Usuario removido."))


@app.route("/admin/logo/upload", methods=["POST"])
def admin_logo_upload():
    arquivo = request.files.get("logo")
    if not arquivo or not (arquivo.filename or "").strip():
        return redirect(url_for("admin", admin_msg="Selecione um arquivo de logo."))

    filename = arquivo.filename.strip()
    ext = os.path.splitext(filename.lower())[1]
    if ext not in LOGO_EXTENSIONS:
        return redirect(url_for("admin", admin_msg="Formato invalido. Use PNG, JPG ou JPEG."))

    image_bytes = arquivo.read()
    if not image_bytes:
        return redirect(url_for("admin", admin_msg="Arquivo de logo vazio."))
    if len(image_bytes) > MAX_LOGO_SIZE_BYTES:
        return redirect(url_for("admin", admin_msg="Logo muito grande. Limite de 12MB."))

    image_bytes, ext, _ = polish_image_bytes(image_bytes, ext, logo=True)

    static_dir = os.path.join(app.root_path, "static")
    os.makedirs(static_dir, exist_ok=True)
    for old_ext in LOGO_EXTENSIONS:
        old_path = os.path.join(static_dir, f"{LOGO_FILE_BASENAME}{old_ext}")
        if os.path.exists(old_path):
            os.remove(old_path)

    target = os.path.join(static_dir, f"{LOGO_FILE_BASENAME}{ext}")
    with open(target, "wb") as fp:
        fp.write(image_bytes)
    return redirect(url_for("admin", admin_msg="Logo atualizada com sucesso."))


@app.route("/admin/importacao-geral", methods=["POST"])
def admin_importacao_geral():
    arquivo = request.files.get("arquivo")
    if not arquivo:
        return redirect(url_for("admin", admin_msg="Selecione um arquivo para importacao em massa."))

    payload, erro = parse_master_upload(arquivo)
    if erro:
        return redirect(url_for("admin", admin_msg=erro))

    stats = import_master_payload(payload)
    resumo = (
        "Importacao geral concluida: "
        f"{stats['obras_novas']} obras, "
        f"{stats['categorias_novas']} categorias, "
        f"{stats['unidades_novas']} unidades, "
        f"{stats['insumos_novos']} insumos novos, "
        f"{stats['insumos_atualizados']} insumos atualizados, "
        f"{stats['vinculos_unidade_novos']} vinculos unidade/insumo, "
        f"{stats['especificacoes_novas']} especificacoes, "
        f"{stats['apropriacoes_novas']} apropriacoes, "
        f"{stats['linhas_invalidas']} linhas invalidas."
    )
    return redirect(url_for("admin", admin_msg=resumo))


@app.route("/admin/importacao-template")
def admin_importacao_template():
    if openpyxl is None:
        return "Instale openpyxl para gerar template XLSX.", 500

    wb = openpyxl.Workbook()
    ws_obras = wb.active
    ws_obras.title = "obras"
    ws_obras.append(["nome"])
    ws_obras.append(["ED. ITALIA"])

    ws_categorias = wb.create_sheet("categorias")
    ws_categorias.append(["nome"])
    ws_categorias.append(["ELETRICA"])
    ws_categorias.append(["HIDRAULICA"])

    ws_unidades = wb.create_sheet("unidades")
    ws_unidades.append(["nome"])
    ws_unidades.append(["UND"])
    ws_unidades.append(["M2"])
    ws_unidades.append(["M3"])

    ws_insumos = wb.create_sheet("insumos")
    ws_insumos.append(["nome", "categoria", "unidades"])
    ws_insumos.append(["CABO FLEXIVEL 2,5MM VERMELHO", "ELETRICA", "UND"])
    ws_insumos.append(["AREIA LAVADA", "HIDRAULICA", "M3"])

    ws_specs = wb.create_sheet("especificacoes")
    ws_specs.append(["insumo", "descricao"])
    ws_specs.append(["CABO FLEXIVEL 2,5MM VERMELHO", "ROLO 100M"])
    ws_specs.append(["AREIA LAVADA", "MEDIA"])

    ws_aprop = wb.create_sheet("apropriacoes")
    ws_aprop.append(["obra", "numero", "nome"])
    ws_aprop.append(["ED. ITALIA", "00.028", "Alvenaria Estrutural"])
    ws_aprop.append(["ED. ITALIA", "00.030", "Estrutura Metalica"])

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return send_file(
        output,
        as_attachment=True,
        download_name="template_importacao_geral.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


@app.route("/admin/obras", methods=["GET", "POST"])
def admin_obras():
    mensagem = request.args.get("msg", "")
    with closing(get_db()) as conn:
        cur = conn.cursor()
        if request.method == "POST":
            nome = request.form.get("nome", "").strip()
            if not nome:
                return redirect(url_for("admin_obras", msg="Informe o nome da obra."))

            key = normalize_text(nome)
            existentes = {
                normalize_text(row["nome"])
                for row in cur.execute("SELECT nome FROM obras").fetchall()
            }
            if key in existentes:
                return redirect(url_for("admin_obras", msg="Obra ja cadastrada."))

            cur.execute("INSERT INTO obras (nome) VALUES (?)", (nome,))
            conn.commit()
            return redirect(url_for("admin_obras", msg="Obra adicionada com sucesso."))
        obras = rows_to_dicts(
            cur.execute("SELECT id, nome FROM obras ORDER BY nome").fetchall()
        )
    return render_template(
        "admin_obras.html",
        title=APP_TITLE,
        obras=obras,
        mensagem=mensagem,
        current_admin="obras",
    )


@app.route("/admin/obras/bulk", methods=["POST"])
def admin_obras_bulk():
    texto = request.form.get("lista", "")
    linhas = [linha.strip() for linha in texto.splitlines() if linha.strip()]
    if not linhas:
        return redirect(url_for("admin_obras", msg="Nenhuma obra informada."))

    with closing(get_db()) as conn:
        cur = conn.cursor()
        existentes = {
            normalize_text(row["nome"])
            for row in cur.execute("SELECT nome FROM obras").fetchall()
        }
        novos = []
        ignoradas = 0
        vistos = set()
        for nome in linhas:
            key = normalize_text(nome)
            if not key or key in existentes or key in vistos:
                ignoradas += 1
                continue
            novos.append((nome,))
            vistos.add(key)
            existentes.add(key)

        if novos:
            cur.executemany("INSERT INTO obras (nome) VALUES (?)", novos)
            conn.commit()
    msg = (
        f"Importacao concluida: {len(novos)} obra(s) nova(s), "
        f"{ignoradas} ignorada(s) por duplicidade/erro."
    )
    return redirect(url_for("admin_obras", msg=msg))


@app.route("/admin/obras/bulk-delete", methods=["POST"])
def admin_obras_bulk_delete():
    obra_ids = parse_int_list(request.form.getlist("obra_ids"))
    if not obra_ids:
        return redirect(url_for("admin_obras", msg="Nenhuma obra selecionada."))

    placeholders = ",".join(["?"] * len(obra_ids))
    with closing(get_db()) as conn:
        conn.execute(
            f"DELETE FROM user_obras WHERE obra_id IN ({placeholders})",
            obra_ids,
        )
        conn.execute(
            f"DELETE FROM apropriacoes WHERE obra_id IN ({placeholders})",
            obra_ids,
        )
        conn.execute(f"DELETE FROM obras WHERE id IN ({placeholders})", obra_ids)
        conn.commit()
    return redirect(url_for("admin_obras", msg=f"{len(obra_ids)} obra(s) removida(s)."))


@app.route("/admin/obras/<int:obra_id>/delete", methods=["POST"])
def admin_obras_delete(obra_id: int):
    with closing(get_db()) as conn:
        conn.execute("DELETE FROM user_obras WHERE obra_id = ?", (obra_id,))
        conn.execute("DELETE FROM apropriacoes WHERE obra_id = ?", (obra_id,))
        conn.execute("DELETE FROM obras WHERE id = ?", (obra_id,))
        conn.commit()
    return redirect_next("admin_obras")


@app.route("/admin/unidades", methods=["GET", "POST"])
def admin_unidades():
    mensagem = request.args.get("msg", "")
    with closing(get_db()) as conn:
        cur = conn.cursor()
        if request.method == "POST":
            nome = (request.form.get("nome") or "").strip()
            if not nome:
                return redirect(url_for("admin_unidades", msg="Informe a unidade."))

            key = normalize_text(nome)
            existentes = {
                normalize_text(row["nome"])
                for row in cur.execute("SELECT nome FROM unidades").fetchall()
            }
            if key in existentes:
                return redirect(url_for("admin_unidades", msg="Unidade ja cadastrada."))

            get_or_create_unidades(cur, [nome])
            conn.commit()
            return redirect(url_for("admin_unidades", msg="Unidade adicionada com sucesso."))

        unidades = rows_to_dicts(
            cur.execute("SELECT id, nome FROM unidades ORDER BY nome").fetchall()
        )
        categorias = rows_to_dicts(
            cur.execute("SELECT id, nome FROM categorias ORDER BY nome").fetchall()
        )
        insumos = rows_to_dicts(
            cur.execute(
                "SELECT id, nome, categoria_id FROM insumos WHERE COALESCE(is_custom, 0) = 0 ORDER BY nome"
            ).fetchall()
        )
        vinculos = rows_to_dicts(
            cur.execute(
                """
                SELECT iu.unidade_id, iu.insumo_id, u.nome as unidade, i.nome as insumo
                FROM insumo_unidades iu
                JOIN unidades u ON u.id = iu.unidade_id
                JOIN insumos i ON i.id = iu.insumo_id
                WHERE COALESCE(i.is_custom, 0) = 0
                ORDER BY u.nome, i.nome
                """
            ).fetchall()
        )

    insumos_por_unidade = {}
    for vinculo in vinculos:
        unidade_id = vinculo["unidade_id"]
        insumos_por_unidade.setdefault(unidade_id, []).append(vinculo["insumo"])

    return render_template(
        "admin_unidades.html",
        title=APP_TITLE,
        unidades=unidades,
        categorias=categorias,
        insumos=insumos,
        vinculos=vinculos,
        insumos_por_unidade=insumos_por_unidade,
        mensagem=mensagem,
        current_admin="unidades",
    )


@app.route("/admin/unidades/bulk", methods=["POST"])
def admin_unidades_bulk():
    texto = request.form.get("lista", "")
    linhas = [linha.strip() for linha in texto.splitlines() if linha.strip()]
    if not linhas:
        return redirect(url_for("admin_unidades", msg="Nenhuma unidade informada."))

    with closing(get_db()) as conn:
        cur = conn.cursor()
        existentes = {
            normalize_text(row["nome"])
            for row in cur.execute("SELECT nome FROM unidades").fetchall()
        }
        novas = []
        vistos = set()
        ignoradas = 0
        for linha in linhas:
            key = normalize_text(linha)
            if not key or key in existentes or key in vistos:
                ignoradas += 1
                continue
            novas.append(linha)
            vistos.add(key)
            existentes.add(key)

        if novas:
            get_or_create_unidades(cur, novas)
            conn.commit()
    msg = (
        f"Importacao concluida: {len(novas)} unidade(s) nova(s), "
        f"{ignoradas} ignorada(s) por duplicidade/erro."
    )
    return redirect(url_for("admin_unidades", msg=msg))


@app.route("/admin/unidades/bulk-delete", methods=["POST"])
def admin_unidades_bulk_delete():
    unidade_ids = parse_int_list(request.form.getlist("unidade_ids"))
    if not unidade_ids:
        return redirect(url_for("admin_unidades", msg="Nenhuma unidade selecionada."))

    placeholders = ",".join(["?"] * len(unidade_ids))
    with closing(get_db()) as conn:
        conn.execute(
            f"DELETE FROM insumo_unidades WHERE unidade_id IN ({placeholders})",
            unidade_ids,
        )
        conn.execute(f"DELETE FROM unidades WHERE id IN ({placeholders})", unidade_ids)
        conn.commit()
    return redirect(
        url_for("admin_unidades", msg=f"{len(unidade_ids)} unidade(s) removida(s).")
    )


@app.route("/admin/unidades/assign", methods=["POST"])
def admin_unidades_assign():
    alvo = (request.form.get("alvo") or "selecionados").strip().lower()
    filtro_insumo = (request.form.get("filtro_insumo") or "").strip()
    categoria_id_raw = (request.form.get("categoria_id") or "").strip()
    categoria_ids = parse_int_list(request.form.getlist("categoria_ids"))
    if not categoria_ids and categoria_id_raw:
        try:
            categoria_ids = [int(categoria_id_raw)]
        except (TypeError, ValueError):
            categoria_ids = []
    insumo_ids = parse_int_list(request.form.getlist("insumo_ids"))
    unidade_ids = parse_int_list(request.form.getlist("unidade_ids"))
    novas_linhas = [
        linha.strip()
        for linha in (request.form.get("novas_unidades") or "").splitlines()
        if linha.strip()
    ]
    with closing(get_db()) as conn:
        cur = conn.cursor()

        if alvo == "todos":
            insumo_ids = parse_int_list(
                [
                    row["id"]
                    for row in cur.execute(
                        "SELECT id FROM insumos WHERE COALESCE(is_custom, 0) = 0"
                    ).fetchall()
                ]
            )
        elif alvo == "filtro":
            if filtro_insumo:
                like = f"%{filtro_insumo.lower()}%"
                insumo_ids = parse_int_list(
                    [
                        row["id"]
                        for row in cur.execute(
                            """
                            SELECT id FROM insumos
                            WHERE lower(nome) LIKE ?
                              AND COALESCE(is_custom, 0) = 0
                            """,
                            (like,),
                        ).fetchall()
                    ]
                )
            else:
                insumo_ids = []
        elif alvo == "categoria":
            if categoria_ids:
                placeholders = ",".join(["?"] * len(categoria_ids))
                insumo_ids = parse_int_list(
                    [
                        row["id"]
                        for row in cur.execute(
                            f"""
                            SELECT id FROM insumos
                            WHERE categoria_id IN ({placeholders})
                              AND COALESCE(is_custom, 0) = 0
                            """,
                            categoria_ids,
                        ).fetchall()
                    ]
                )
            else:
                insumo_ids = []

        if not insumo_ids:
            return redirect_next("admin_unidades")

        novas_unidade_ids = get_or_create_unidades(cur, novas_linhas)
        unidade_ids = sorted(set(unidade_ids + novas_unidade_ids))
        if not unidade_ids:
            conn.commit()
            return redirect_next("admin_unidades")

        pares = [
            (insumo_id, unidade_id)
            for insumo_id in insumo_ids
            for unidade_id in unidade_ids
        ]
        cur.executemany(
            """
            INSERT OR IGNORE INTO insumo_unidades (insumo_id, unidade_id)
            VALUES (?, ?)
            """,
            pares,
        )
        conn.commit()
    return redirect_next("admin_unidades")


@app.route("/admin/unidades/<int:unidade_id>/delete", methods=["POST"])
def admin_unidades_delete(unidade_id: int):
    with closing(get_db()) as conn:
        conn.execute("DELETE FROM insumo_unidades WHERE unidade_id = ?", (unidade_id,))
        conn.execute("DELETE FROM unidades WHERE id = ?", (unidade_id,))
        conn.commit()
    return redirect_next("admin_unidades")


@app.route("/admin/unidades/vinculo/delete", methods=["POST"])
def admin_unidades_vinculo_delete():
    try:
        insumo_id = int(request.form.get("insumo_id", "0"))
        unidade_id = int(request.form.get("unidade_id", "0"))
    except (TypeError, ValueError):
        return redirect_next("admin_unidades")

    if not insumo_id or not unidade_id:
        return redirect_next("admin_unidades")

    with closing(get_db()) as conn:
        conn.execute(
            "DELETE FROM insumo_unidades WHERE insumo_id = ? AND unidade_id = ?",
            (insumo_id, unidade_id),
        )
        conn.commit()
    return redirect_next("admin_unidades")


@app.route("/admin/insumos", methods=["GET", "POST"])
def admin_insumos():
    mensagem = request.args.get("msg", "")
    with closing(get_db()) as conn:
        cur = conn.cursor()
        if request.method == "POST":
            nome = request.form.get("nome", "").strip()
            categoria_id = request.form.get("categoria_id") or None
            if not nome:
                return redirect(url_for("admin_insumos", msg="Informe o nome do insumo."))

            key = normalize_text(nome)
            existentes = {
                normalize_text(row["nome"])
                for row in cur.execute(
                    "SELECT nome FROM insumos WHERE COALESCE(is_custom, 0) = 0"
                ).fetchall()
            }
            if key in existentes:
                return redirect(url_for("admin_insumos", msg="Insumo ja cadastrado."))

            cur.execute(
                "INSERT INTO insumos (nome, categoria_id) VALUES (?, ?)",
                (nome, categoria_id),
            )
            conn.commit()
            return redirect(url_for("admin_insumos", msg="Insumo adicionado com sucesso."))
        categorias = rows_to_dicts(
            cur.execute("SELECT id, nome FROM categorias ORDER BY nome").fetchall()
        )
        unidades = rows_to_dicts(
            cur.execute("SELECT id, nome FROM unidades ORDER BY nome").fetchall()
        )
        insumos = rows_to_dicts(
            cur.execute(
                """
                SELECT
                    i.id,
                    i.nome,
                    i.categoria_id,
                    c.nome as categoria,
                    (
                        SELECT GROUP_CONCAT(u2.nome, ', ')
                        FROM insumo_unidades iu2
                        JOIN unidades u2 ON u2.id = iu2.unidade_id
                        WHERE iu2.insumo_id = i.id
                    ) as unidades
                FROM insumos i
                LEFT JOIN categorias c ON c.id = i.categoria_id
                WHERE COALESCE(i.is_custom, 0) = 0
                ORDER BY i.nome
                """
            ).fetchall()
        )
    return render_template(
        "admin_insumos.html",
        title=APP_TITLE,
        insumos=insumos,
        categorias=categorias,
        unidades=unidades,
        mensagem=mensagem,
        current_admin="insumos",
    )


@app.route("/admin/insumos/upload", methods=["POST"])
def admin_insumos_upload():
    arquivo = request.files.get("arquivo")
    if not arquivo:
        return redirect(url_for("admin_insumos", msg="Selecione um arquivo CSV ou XLSX."))

    parsed, erro = parse_insumos_upload(arquivo)
    if erro:
        return redirect(url_for("admin_insumos", msg=erro))

    stats = import_master_payload(
        {
            "obras": [],
            "categorias": [],
            "unidades": [],
            "insumos": parsed["insumos"],
            "especificacoes": parsed["especificacoes"],
            "apropriacoes": [],
        }
    )
    resumo = (
        "Importacao de insumos concluida: "
        f"{stats['insumos_novos']} novos, "
        f"{stats['insumos_atualizados']} atualizados, "
        f"{stats['categorias_novas']} categorias novas, "
        f"{stats['unidades_novas']} unidades novas, "
        f"{stats['vinculos_unidade_novos']} vinculos unidade/insumo, "
        f"{stats['especificacoes_novas']} especificacoes, "
        f"{stats['linhas_invalidas']} linhas invalidas."
    )
    return redirect(url_for("admin_insumos", msg=resumo))


@app.route("/admin/insumos/template")
def admin_insumos_template():
    if openpyxl is None:
        output = BytesIO()
        output.write(
            (
                "nome,categoria,unidades,especificacoes\n"
                "CABO FLEXIVEL 2,5MM VERMELHO,ELETRICA,UND|M,ROLO 100M|ANTI-CHAMA\n"
                "AREIA MEDIA,ALVENARIA,M3,SEM IMPUREZAS\n"
            ).encode("utf-8-sig")
        )
        output.seek(0)
        return send_file(
            output,
            as_attachment=True,
            download_name="template_insumos.csv",
            mimetype="text/csv",
        )

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "insumos"
    ws.append(["nome", "categoria", "unidades", "especificacoes"])
    ws.append(
        [
            "CABO FLEXIVEL 2,5MM VERMELHO",
            "ELETRICA",
            "UND|M",
            "ROLO 100M|ANTI-CHAMA",
        ]
    )
    ws.append(["AREIA MEDIA", "ALVENARIA", "M3", "SEM IMPUREZAS"])

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return send_file(
        output,
        as_attachment=True,
        download_name="template_insumos.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


@app.route("/admin/insumos/<int:insumo_id>/edit", methods=["GET", "POST"])
def admin_insumos_edit(insumo_id: int):
    mensagem = request.args.get("msg", "")
    with closing(get_db()) as conn:
        cur = conn.cursor()
        atual = cur.execute(
            "SELECT id, nome, categoria_id FROM insumos WHERE id = ?",
            (insumo_id,),
        ).fetchone()
        if atual is None:
            return redirect(url_for("admin_insumos"))

        if request.method == "POST":
            nome = request.form.get("nome", "").strip()
            categoria_id = request.form.get("categoria_id") or None
            unidade_ids = parse_int_list(request.form.getlist("unidade_ids"))
            novas_linhas = [
                linha.strip()
                for linha in (request.form.get("novas_unidades") or "").splitlines()
                if linha.strip()
            ]
            novas_unidade_ids = get_or_create_unidades(cur, novas_linhas)
            unidade_ids = sorted(set(unidade_ids + novas_unidade_ids))

            if not nome:
                return redirect(url_for("admin_insumos", msg="Informe o nome do insumo."))

            key = normalize_text(nome)
            for row in cur.execute(
                "SELECT id, nome FROM insumos WHERE id <> ?",
                (insumo_id,),
            ).fetchall():
                if normalize_text(row["nome"]) == key:
                    return redirect(
                        url_for("admin_insumos", msg="Ja existe um insumo com esse nome.")
                    )
            cur.execute(
                """
                UPDATE insumos
                SET nome = ?, categoria_id = ?
                WHERE id = ?
                """,
                (nome, categoria_id, insumo_id),
            )
            if (request.form.get("gerenciar_unidades") or "") == "1":
                cur.execute(
                    "DELETE FROM insumo_unidades WHERE insumo_id = ?",
                    (insumo_id,),
                )
                if unidade_ids:
                    cur.executemany(
                        """
                        INSERT OR IGNORE INTO insumo_unidades (insumo_id, unidade_id)
                        VALUES (?, ?)
                        """,
                        [(insumo_id, unidade_id) for unidade_id in unidade_ids],
                    )
            conn.commit()
            return redirect(url_for("admin_insumos", msg="Insumo atualizado com sucesso."))

        insumo = dict(atual)
        categorias = rows_to_dicts(
            cur.execute("SELECT id, nome FROM categorias ORDER BY nome").fetchall()
        )
        unidades = rows_to_dicts(
            cur.execute("SELECT id, nome FROM unidades ORDER BY nome").fetchall()
        )
        insumo_unidade_ids = [
            row["unidade_id"]
            for row in cur.execute(
                "SELECT unidade_id FROM insumo_unidades WHERE insumo_id = ?",
                (insumo_id,),
            ).fetchall()
        ]
    return render_template(
        "admin_insumos_edit.html",
        title=APP_TITLE,
        insumo=insumo,
        categorias=categorias,
        unidades=unidades,
        insumo_unidade_ids=insumo_unidade_ids,
        mensagem=mensagem,
        current_admin="insumos",
    )


@app.route("/admin/insumos/bulk", methods=["POST"])
def admin_insumos_bulk():
    texto = request.form.get("lista", "")
    categoria_id = request.form.get("categoria_id") or None
    linhas = [linha.strip() for linha in texto.splitlines() if linha.strip()]
    if not linhas:
        return redirect(url_for("admin_insumos", msg="Nenhum insumo informado."))

    with closing(get_db()) as conn:
        cur = conn.cursor()
        existentes = {
            normalize_text(row["nome"])
            for row in cur.execute("SELECT nome FROM insumos").fetchall()
        }
        novos = []
        ignorados = 0
        vistos = set()
        for nome in linhas:
            key = normalize_text(nome)
            if not key or key in existentes or key in vistos:
                ignorados += 1
                continue
            novos.append((nome, categoria_id))
            vistos.add(key)
            existentes.add(key)

        if novos:
            cur.executemany(
                "INSERT INTO insumos (nome, categoria_id) VALUES (?, ?)", novos
            )
            conn.commit()
    msg = (
        f"Importacao concluida: {len(novos)} insumo(s) novo(s), "
        f"{ignorados} ignorado(s) por duplicidade/erro."
    )
    return redirect(url_for("admin_insumos", msg=msg))


@app.route("/admin/insumos/bulk-delete", methods=["POST"])
def admin_insumos_bulk_delete():
    raw_ids = request.form.getlist("insumo_ids")
    insumo_ids = []
    for raw_id in raw_ids:
        try:
            insumo_ids.append(int(raw_id))
        except (TypeError, ValueError):
            continue

    if not insumo_ids:
        return redirect_next("admin_insumos")

    unique_ids = sorted(set(insumo_ids))
    placeholders = ",".join(["?"] * len(unique_ids))
    with closing(get_db()) as conn:
        fotos = rows_to_dicts(
            conn.execute(
                f"SELECT foto_path FROM insumos WHERE id IN ({placeholders})",
                unique_ids,
            ).fetchall()
        )
        conn.execute(
            f"DELETE FROM insumo_unidades WHERE insumo_id IN ({placeholders})",
            unique_ids,
        )
        conn.execute(
            f"DELETE FROM insumos WHERE id IN ({placeholders})",
            unique_ids,
        )
        conn.commit()
    for row in fotos:
        remove_insumo_image(row.get("foto_path", ""))
    return redirect_next("admin_insumos")


@app.route("/admin/insumos/<int:insumo_id>/delete", methods=["POST"])
def admin_insumos_delete(insumo_id: int):
    with closing(get_db()) as conn:
        row = conn.execute(
            "SELECT foto_path FROM insumos WHERE id = ?", (insumo_id,)
        ).fetchone()
        conn.execute("DELETE FROM insumo_unidades WHERE insumo_id = ?", (insumo_id,))
        conn.execute("DELETE FROM insumos WHERE id = ?", (insumo_id,))
        conn.commit()
    if row is not None:
        remove_insumo_image(row["foto_path"] or "")
    return redirect_next("admin_insumos")


@app.route("/admin/especificacoes", methods=["GET", "POST"])
def admin_especificacoes():
    import_result = request.args.get("import_result", "")
    with closing(get_db()) as conn:
        cur = conn.cursor()
        if request.method == "POST":
            insumo_id = request.form.get("insumo_id")
            nome = request.form.get("nome", "").strip()
            if not insumo_id or not nome:
                return redirect(
                    url_for(
                        "admin_especificacoes",
                        import_result="Informe insumo e especificacao.",
                    )
                )

            key = normalize_text(nome)
            existentes = {
                normalize_text(row["nome"])
                for row in cur.execute(
                    "SELECT nome FROM especificacoes WHERE insumo_id = ?",
                    (insumo_id,),
                ).fetchall()
            }
            if key in existentes:
                return redirect(
                    url_for(
                        "admin_especificacoes",
                        import_result="Essa especificacao ja existe para o insumo.",
                    )
                )

            cur.execute(
                "INSERT INTO especificacoes (insumo_id, nome) VALUES (?, ?)",
                (insumo_id, nome),
            )
            conn.commit()
            return redirect(
                url_for(
                    "admin_especificacoes",
                    import_result="Especificacao adicionada com sucesso.",
                )
            )
        insumos = rows_to_dicts(
            cur.execute(
                "SELECT id, nome FROM insumos WHERE COALESCE(is_custom, 0) = 0 ORDER BY nome"
            ).fetchall()
        )
        specs = rows_to_dicts(
            cur.execute(
                """
                SELECT e.id, e.nome, i.nome as insumo
                FROM especificacoes e
                JOIN insumos i ON i.id = e.insumo_id
                WHERE COALESCE(i.is_custom, 0) = 0
                ORDER BY i.nome, e.nome
                """
            ).fetchall()
        )
    return render_template(
        "admin_especificacoes.html",
        title=APP_TITLE,
        insumos=insumos,
        specs=specs,
        import_result=import_result,
        current_admin="especificacoes",
    )


@app.route("/admin/especificacoes/bulk", methods=["POST"])
def admin_especificacoes_bulk():
    insumo_id = request.form.get("insumo_id")
    texto = request.form.get("lista", "")
    linhas = [linha.strip() for linha in texto.splitlines() if linha.strip()]
    if not insumo_id or not linhas:
        return redirect(
            url_for(
                "admin_especificacoes",
                import_result="Informe insumo e ao menos uma especificacao.",
            )
        )

    with closing(get_db()) as conn:
        cur = conn.cursor()
        existentes = {
            normalize_text(row["nome"])
            for row in cur.execute(
                "SELECT nome FROM especificacoes WHERE insumo_id = ?", (insumo_id,)
            ).fetchall()
        }
        novas = []
        ignoradas = 0
        vistos = set()
        for nome in linhas:
            key = normalize_text(nome)
            if not key or key in existentes or key in vistos:
                ignoradas += 1
                continue
            novas.append((insumo_id, nome))
            vistos.add(key)
            existentes.add(key)

        if novas:
            cur.executemany(
                "INSERT INTO especificacoes (insumo_id, nome) VALUES (?, ?)", novas
            )
            conn.commit()
    msg = (
        f"Importacao concluida: {len(novas)} especificacao(oes) nova(s), "
        f"{ignoradas} ignorada(s) por duplicidade/erro."
    )
    return redirect(url_for("admin_especificacoes", import_result=msg))


@app.route("/admin/especificacoes/bulk-delete", methods=["POST"])
def admin_especificacoes_bulk_delete():
    spec_ids = parse_int_list(request.form.getlist("spec_ids"))
    if not spec_ids:
        return redirect(
            url_for("admin_especificacoes", import_result="Nenhuma especificacao selecionada.")
        )

    placeholders = ",".join(["?"] * len(spec_ids))
    with closing(get_db()) as conn:
        conn.execute(f"DELETE FROM especificacoes WHERE id IN ({placeholders})", spec_ids)
        conn.commit()
    return redirect(
        url_for(
            "admin_especificacoes",
            import_result=f"{len(spec_ids)} especificacao(oes) removida(s).",
        )
    )


@app.route("/admin/especificacoes/import", methods=["POST"])
def admin_especificacoes_import():
    arquivo = request.files.get("arquivo")
    criar_insumos = (request.form.get("criar_insumos") or "") == "1"
    next_url = (request.form.get("next") or "").strip()
    target_endpoint = (
        "admin"
        if next_url.startswith("/admin") and not next_url.startswith("//")
        else "admin_especificacoes"
    )
    if not arquivo:
        target = url_for(target_endpoint, import_result="Sem arquivo.")
        if target_endpoint == "admin":
            target += "#especificacoes"
        return redirect(target)

    rows, erro = parse_specs_upload(arquivo)
    if erro:
        target = url_for(target_endpoint, import_result=erro)
        if target_endpoint == "admin":
            target += "#especificacoes"
        return redirect(target)

    with closing(get_db()) as conn:
        cur = conn.cursor()
        insumos_rows = cur.execute(
            "SELECT id, nome FROM insumos WHERE COALESCE(is_custom, 0) = 0"
        ).fetchall()
        insumo_map = {normalize_text(row["nome"]): row["id"] for row in insumos_rows}

        created_insumos = 0
        inserted_specs = 0
        skipped_specs = 0
        missing_insumos = 0

        specs_by_insumo = {}
        for row in cur.execute("SELECT insumo_id, nome FROM especificacoes").fetchall():
            key = row["insumo_id"]
            if key not in specs_by_insumo:
                specs_by_insumo[key] = set()
            specs_by_insumo[key].add(normalize_text(row["nome"]))

        for insumo_nome, descricao in rows:
            insumo_key = normalize_text(insumo_nome)
            insumo_id = insumo_map.get(insumo_key)
            if insumo_id is None:
                if not criar_insumos:
                    missing_insumos += 1
                    continue
                cur.execute("INSERT INTO insumos (nome) VALUES (?)", (insumo_nome.strip(),))
                insumo_id = cur.lastrowid
                insumo_map[insumo_key] = insumo_id
                specs_by_insumo[insumo_id] = set()
                created_insumos += 1

            desc_key = normalize_text(descricao)
            existing = specs_by_insumo.setdefault(insumo_id, set())
            if desc_key in existing:
                skipped_specs += 1
                continue

            cur.execute(
                "INSERT INTO especificacoes (insumo_id, nome) VALUES (?, ?)",
                (insumo_id, descricao.strip()),
            )
            existing.add(desc_key)
            inserted_specs += 1

        conn.commit()

    resumo = (
        f"Importacao concluida: {inserted_specs} descricoes novas, "
        f"{skipped_specs} ja existiam, {created_insumos} insumos criados, "
        f"{missing_insumos} linhas sem insumo cadastrado."
    )
    target = url_for(target_endpoint, import_result=resumo)
    if target_endpoint == "admin":
        target += "#especificacoes"
    return redirect(target)


@app.route("/admin/especificacoes/<int:spec_id>/delete", methods=["POST"])
def admin_especificacoes_delete(spec_id: int):
    with closing(get_db()) as conn:
        conn.execute("DELETE FROM especificacoes WHERE id = ?", (spec_id,))
        conn.commit()
    return redirect_next("admin_especificacoes")


@app.route("/admin/categorias", methods=["GET", "POST"])
def admin_categorias():
    mensagem = request.args.get("msg", "")
    with closing(get_db()) as conn:
        cur = conn.cursor()
        if request.method == "POST":
            nome = request.form.get("nome", "").strip()
            if not nome:
                return redirect(url_for("admin_categorias", msg="Informe o nome da categoria."))

            key = normalize_text(nome)
            existentes = {
                normalize_text(row["nome"])
                for row in cur.execute("SELECT nome FROM categorias").fetchall()
            }
            if key in existentes:
                return redirect(url_for("admin_categorias", msg="Categoria ja cadastrada."))

            cur.execute("INSERT INTO categorias (nome) VALUES (?)", (nome,))
            conn.commit()
            return redirect(url_for("admin_categorias", msg="Categoria adicionada com sucesso."))
        categorias = rows_to_dicts(
            cur.execute("SELECT id, nome FROM categorias ORDER BY nome").fetchall()
        )
    return render_template(
        "admin_categorias.html",
        title=APP_TITLE,
        categorias=categorias,
        mensagem=mensagem,
        current_admin="categorias",
    )


@app.route("/admin/categorias/bulk", methods=["POST"])
def admin_categorias_bulk():
    texto = request.form.get("lista", "")
    linhas = [linha.strip() for linha in texto.splitlines() if linha.strip()]
    if not linhas:
        return redirect(url_for("admin_categorias", msg="Nenhuma categoria informada."))

    with closing(get_db()) as conn:
        cur = conn.cursor()
        existentes = {
            normalize_text(row["nome"])
            for row in cur.execute("SELECT nome FROM categorias").fetchall()
        }
        novos = []
        ignoradas = 0
        vistos = set()
        for nome in linhas:
            key = normalize_text(nome)
            if not key or key in existentes or key in vistos:
                ignoradas += 1
                continue
            novos.append((nome,))
            vistos.add(key)
            existentes.add(key)

        if novos:
            cur.executemany("INSERT INTO categorias (nome) VALUES (?)", novos)
            conn.commit()
    msg = (
        f"Importacao concluida: {len(novos)} categoria(s) nova(s), "
        f"{ignoradas} ignorada(s) por duplicidade/erro."
    )
    return redirect(url_for("admin_categorias", msg=msg))


@app.route("/admin/categorias/bulk-delete", methods=["POST"])
def admin_categorias_bulk_delete():
    categoria_ids = parse_int_list(request.form.getlist("categoria_ids"))
    if not categoria_ids:
        return redirect(url_for("admin_categorias", msg="Nenhuma categoria selecionada."))

    placeholders = ",".join(["?"] * len(categoria_ids))
    with closing(get_db()) as conn:
        conn.execute(
            f"UPDATE insumos SET categoria_id = NULL WHERE categoria_id IN ({placeholders})",
            categoria_ids,
        )
        conn.execute(f"DELETE FROM categorias WHERE id IN ({placeholders})", categoria_ids)
        conn.commit()
    return redirect(
        url_for("admin_categorias", msg=f"{len(categoria_ids)} categoria(s) removida(s).")
    )


@app.route("/admin/categorias/<int:categoria_id>/delete", methods=["POST"])
def admin_categorias_delete(categoria_id: int):
    with closing(get_db()) as conn:
        conn.execute("DELETE FROM categorias WHERE id = ?", (categoria_id,))
        conn.commit()
    return redirect_next("admin_categorias")


@app.route("/admin/apropriacoes", methods=["GET", "POST"])
def admin_apropriacoes():
    mensagem = request.args.get("msg", "")
    with closing(get_db()) as conn:
        cur = conn.cursor()
        if request.method == "POST":
            obra_id = request.form.get("obra_id")
            numero = request.form.get("numero", "").strip().replace(",", ".")
            nome = request.form.get("nome", "").strip()
            if not obra_id or not is_valid_apropriacao(numero) or not nome:
                return redirect(
                    url_for(
                        "admin_apropriacoes",
                        msg="Informe obra, codigo valido e nome da apropriacao.",
                    )
                )

            duplicado = cur.execute(
                "SELECT 1 FROM apropriacoes WHERE obra_id = ? AND numero = ?",
                (obra_id, numero),
            ).fetchone()
            if duplicado:
                return redirect(
                    url_for(
                        "admin_apropriacoes",
                        msg="Apropriacao ja cadastrada para esta obra.",
                    )
                )

            cur.execute(
                "INSERT INTO apropriacoes (obra_id, numero, nome) VALUES (?, ?, ?)",
                (obra_id, numero, nome),
            )
            conn.commit()
            return redirect(
                url_for("admin_apropriacoes", msg="Apropriacao adicionada com sucesso.")
            )

        obras = rows_to_dicts(
            cur.execute("SELECT id, nome FROM obras ORDER BY nome").fetchall()
        )
        apropriacoes = rows_to_dicts(
            cur.execute(
                """
                SELECT
                    a.id,
                    a.obra_id,
                    a.numero,
                    TRIM(COALESCE(a.nome, '')) AS nome,
                    o.nome as obra
                FROM apropriacoes a
                JOIN obras o ON o.id = a.obra_id
                ORDER BY o.nome, a.numero
                """
            ).fetchall()
        )
    apropriacoes_by_obra_id = {}
    for item in apropriacoes:
        obra_id = item.get("obra_id")
        if obra_id not in apropriacoes_by_obra_id:
            apropriacoes_by_obra_id[obra_id] = []
        apropriacoes_by_obra_id[obra_id].append(item)

    apropriacoes_por_obra = []
    for obra in obras:
        itens = apropriacoes_by_obra_id.get(obra["id"], [])
        apropriacoes_por_obra.append(
            {
                "obra_id": obra["id"],
                "obra_nome": obra["nome"],
                "itens": itens,
                "total": len(itens),
            }
        )
    return render_template(
        "admin_apropriacoes.html",
        title=APP_TITLE,
        obras=obras,
        apropriacoes=apropriacoes,
        apropriacoes_por_obra=apropriacoes_por_obra,
        mensagem=mensagem,
        current_admin="apropriacoes",
    )


@app.route("/admin/apropriacoes/bulk", methods=["POST"])
def admin_apropriacoes_bulk():
    obra_id = request.form.get("obra_id")
    texto = request.form.get("lista", "") or ""
    linhas = [linha.strip() for linha in texto.splitlines() if linha.strip()]
    if not obra_id or not linhas:
        return redirect(
            url_for(
                "admin_apropriacoes",
                msg="Informe obra e ao menos uma apropriacao em formato codigo|nome (ou nome|codigo).",
            )
        )

    with closing(get_db()) as conn:
        cur = conn.cursor()
        existentes = {
            row["numero"]: (row["nome"] or "").strip()
            for row in cur.execute(
                "SELECT numero, COALESCE(nome, '') AS nome FROM apropriacoes WHERE obra_id = ?",
                (obra_id,),
            ).fetchall()
        }
        novos = []
        ignoradas = 0
        atualizadas = 0
        for numero in linhas:
            codigo, nome = parse_apropriacao_entry(numero)
            if not codigo or not nome or not is_valid_apropriacao(codigo):
                ignoradas += 1
                continue
            if codigo in existentes:
                nome_existente = existentes[codigo]
                if normalize_text(nome_existente) != normalize_text(nome):
                    cur.execute(
                        "UPDATE apropriacoes SET nome = ? WHERE obra_id = ? AND numero = ?",
                        (nome, obra_id, codigo),
                    )
                    existentes[codigo] = nome
                    atualizadas += 1
                else:
                    ignoradas += 1
                continue
            novos.append((obra_id, codigo, nome))
            existentes[codigo] = nome

        if novos:
            cur.executemany(
                "INSERT INTO apropriacoes (obra_id, numero, nome) VALUES (?, ?, ?)",
                novos,
            )
        if novos or atualizadas:
            conn.commit()
    msg = (
        f"Importacao concluida: {len(novos)} apropriacao(oes) nova(s), "
        f"{atualizadas} atualizada(s), "
        f"{ignoradas} ignorada(s) por duplicidade/erro."
    )
    return redirect(url_for("admin_apropriacoes", msg=msg))


@app.route("/admin/apropriacoes/bulk-delete", methods=["POST"])
def admin_apropriacoes_bulk_delete():
    apropriacao_ids = parse_int_list(request.form.getlist("apropriacao_ids"))
    if not apropriacao_ids:
        return redirect(
            url_for("admin_apropriacoes", msg="Nenhuma apropriacao selecionada.")
        )

    placeholders = ",".join(["?"] * len(apropriacao_ids))
    with closing(get_db()) as conn:
        conn.execute(
            f"DELETE FROM apropriacoes WHERE id IN ({placeholders})",
            apropriacao_ids,
        )
        conn.commit()
    return redirect(
        url_for(
            "admin_apropriacoes",
            msg=f"{len(apropriacao_ids)} apropriacao(oes) removida(s).",
        )
    )


@app.route("/admin/apropriacoes/<int:apropriacao_id>/delete", methods=["POST"])
def admin_apropriacoes_delete(apropriacao_id: int):
    with closing(get_db()) as conn:
        conn.execute("DELETE FROM apropriacoes WHERE id = ?", (apropriacao_id,))
        conn.commit()
    return redirect_next("admin_apropriacoes")


# Garante as tabelas no import (ex.: gunicorn/PythonAnywhere).
init_db()


if __name__ == "__main__":
    debug_mode = (os.environ.get("FLASK_DEBUG") or "").strip() == "1"
    app.run(debug=debug_mode, use_reloader=debug_mode)
